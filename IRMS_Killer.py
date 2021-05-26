import base64
import collections
import datetime
import concurrent.futures
import copy
import Crypto.Cipher.AES
import csv
import json
import math
import lxml
import openpyxl
import re
import requests
import urllib

'''
数据完整性验(http返回状态统计,数据数量统计,SFB,OCS)
小区状态统计(ODM,)
数据测存储与验证
重新主线流程
'''

File_Name = ['平舆小黑']

def Swimming_Pool(Para_Functional_Function,Para_Some_Iterable_Obj):
    with concurrent.futures.ThreadPoolExecutor(max_workers=10) as Pool_Executor:
        Pool_Executor.map(Para_Functional_Function,(Para_Some_Iterable_Obj))

def Generate_Local_Data(Para_File_Name):
    '''读取本地数据,在不查询的前提下填充几个List'''

    '''读取并整理7013表,生成List_7013'''
    global List_7013
    List_7013 = []
    List_Original_7013 = []
    List_Original_Num = []
    with open(Para_File_Name+'.csv') as file_csv:
        reader_obj = csv.reader(file_csv)
        List_Original_7013 = list(reader_obj)
    List_7013_Use = ['所属地市', '所属区县', '资管中文名称', '安装位置', '项目编号', '任务名称', '主用OLT', '主用OLT的PON端口', '上级POS名称', '上级POS端口', '级别', '中文名称']
    for column_num, column_name in enumerate(List_7013_Use):
        for column_num_2, column_name_2 in enumerate(List_Original_7013[0]):
            if column_name == column_name_2:
                List_Original_Num.append(column_num_2)
    for each_original_7013 in List_Original_7013:
        each_7013 = []
        for index_num in List_Original_Num:
            each_7013.append(each_original_7013[index_num])
        List_7013.append(each_7013)
    Task_Name_ID_List = List_7013[1][5].split('-')

    '''读取并整理Sheet_Template,生成List_Template'''
    WB_obj = openpyxl.load_workbook(Para_File_Name+'.xlsx') 

    WS_obj = WB_obj['Info']
    List_Template = []
    cell_range = WS_obj['A21': 'U31']
    for row_data in cell_range:
        List_Temp_1 = []
        for cell in row_data:
            List_Temp_1.append(cell.value)
        List_Template.append(List_Temp_1)
    for list_num in List_Template:
        if List_7013[1][1] == list_num[0]:
            List_Template_Selected = copy.deepcopy(list_num)

    global Username_Create,Password_Create,Username_Run,Password_Run
    Username_Create = List_Template_Selected[17]
    Password_Create = List_Template_Selected[18]
    Username_Run = List_Template_Selected[19]
    Password_Run = List_Template_Selected[20]

    '''读取并整理Sheet_Info,生成List_Box_Type和其他参数'''
    Longitude_Start     = WS_obj['B2'].value
    Latitude_Start      = WS_obj['B3'].value
    Horizontal_Density  = WS_obj['B4'].value
    Vertical_Density    = WS_obj['B5'].value
    Anchor_Point_Buttom = WS_obj['B6'].value
    Anchor_Point_Right  = WS_obj['B7'].value

    global P0_Data_Check,P1_Push_Box,P2_Generate_Support_Segment,P3_Generate_Cable_Segment,P4_Cable_Lay,P5_Generate_ODM,P6_Generate_Tray,P7_Termination,P8_Direct_Melt,P9_Generate_Optical_Circuit, P10_Transmission_Design,P11_Termination,P12_Update_1_Fix_OCS
    P0_Data_Check               = WS_obj['E2'].value
    P1_Push_Box                 = WS_obj['E3'].value
    P2_Generate_Support_Segment = WS_obj['E4'].value
    P3_Generate_Cable_Segment   = WS_obj['E5'].value
    P4_Cable_Lay                = WS_obj['E6'].value
    P5_Generate_ODM             = WS_obj['E7'].value
    P6_Generate_Tray            = WS_obj['E8'].value
    P7_Termination              = WS_obj['E9'].value
    P8_Direct_Melt              = WS_obj['E10'].value
    P9_Generate_Optical_Circuit = WS_obj['E11'].value
    P10_Transmission_Design     = WS_obj['E12'].value
    P11_Termination             = WS_obj['E13'].value
    P12_Update_1_Fix_OCS        = WS_obj['E14'].value

    if (P9_Generate_Optical_Circuit and P10_Transmission_Design):
        print('光路申请与传输设计不能同时进行')
        exit()
    if (P11_Termination and P10_Transmission_Design):
        print('传输设计与备芯上架不能同时进行')
        exit()
    if (P9_Generate_Optical_Circuit and P11_Termination):
        print('光路申请与备芯上架不能同时进行')
        exit()

    # 根据照片修改数据
    global List_Modify_For_Photo
    List_Modify_For_Photo = []
    box_num = 0
    while WS_obj[('H' + str(2 + box_num))].value:
        List_Modify_For_Photo.append([WS_obj[('H' + str(2 + box_num))].value, WS_obj[('I' + str(2 + box_num))].value, WS_obj[('J' + str(2 + box_num))].value])
        box_num += 1

    # 箱体类型ID
    List_Box_Type = [['guangfenxianxiang', 9204, 9115], ['guangjiaojiexiang', 9203, 9115]]

    '''读取并整理Sheet_Box_Topology,生成List_Box_Data'''
    WS_obj = WB_obj['BOX_Topology']
    List_Box_Name = []
    for row_num in range(1,201):
        for column_num in range(1,201):
            v = WS_obj.cell(row_num, column_num)
            if v.value == 'empty':
                continue
            if v.value == None:
                break
            List_Box_Name.append(list([v.value, row_num, column_num]))

    global List_Box_Data
    List_Box_Data = []
    for box_num in range(len(List_Box_Name)):
        List_Box_Data.append(dict({'Box_Name':List_Box_Name[box_num][0]}))
        List_Box_Data[box_num]['Longitude'] = Longitude_Start + (List_Box_Name[box_num][2] - 1) * Horizontal_Density
        List_Box_Data[box_num]['Latitude'] = Latitude_Start - (List_Box_Name[box_num][1] - 1) * Vertical_Density
        if (List_Box_Data[box_num]['Box_Name'].find('GJ') != -1) or (List_Box_Data[box_num]['Box_Name'].find('gj') != -1):
            List_Box_Data[box_num]['Box_Type'] = 'guangjiaojiexiang'
        else:
            List_Box_Data[box_num]['Box_Type'] = 'guangfenxianxiang'
        for row_data in range(len(List_Box_Type)):
            if List_Box_Data[box_num]['Box_Type'] == List_Box_Type[row_data][0]:
                List_Box_Data[box_num]['Box_Type_ID'] = List_Box_Type[row_data][1]
                List_Box_Data[box_num]['ResPoint_Type_ID'] = List_Box_Type[row_data][2]
        List_Box_Data[box_num]['City_ID'] = List_Template_Selected[16]
        List_Box_Data[box_num]['County_ID'] = List_Template_Selected[1]

    '''处理锚点'''
    if Anchor_Point_Buttom:
        Move_Up = List_Box_Name[-1][1]
        for each_box_data in List_Box_Data:
            each_box_data['Latitude'] = each_box_data['Latitude'] + Vertical_Density * (Move_Up - 1)
    if Anchor_Point_Right:
        Move_Left = List_Box_Name[-1][2]
        for each_box_data in List_Box_Data:
            each_box_data['Longitude'] = each_box_data['Longitude'] - Horizontal_Density * (Move_Left - 1)

    if P0_Data_Check or P2_Generate_Support_Segment or P3_Generate_Cable_Segment or P4_Cable_Lay or P5_Generate_ODM or P6_Generate_Tray or P7_Termination or P8_Direct_Melt or P9_Generate_Optical_Circuit or P10_Transmission_Design or P11_Termination or P12_Update_1_Fix_OCS:
        '''读取并整理Sheet_OCS_List,生成List_CS_Data'''
        WS_obj = WB_obj['OCS_List']
        global List_CS_Data
        List_CS_Data = []
        for row_num in range(1,501):
            OCS_A_Box_Name = WS_obj.cell(row_num, 1).value
            OCS_Z_Box_Name = WS_obj.cell(row_num, 2).value
            OCS_Width = WS_obj.cell(row_num, 3).value
            if OCS_A_Box_Name == None:
                break
            List_CS_Data.append(dict({'A_Box_Name': OCS_A_Box_Name}))
            List_CS_Data[row_num - 1]['Z_Box_Name'] = OCS_Z_Box_Name
            List_CS_Data[row_num - 1]['Width'] = OCS_Width
        
        #Length_Prepare
        Horizontal_Metre = 111.11 * 1000 * math.cos(Latitude_Start * math.pi / 180)
        Vertical_Metre = 111.11 * 1000
        
        for dic_num_in_osc in List_CS_Data:
            for dic_num_in_box in List_Box_Data:
                if dic_num_in_osc['A_Box_Name'] == dic_num_in_box['Box_Name']:
                    dic_num_in_osc['A_Box_Type_ID'] = dic_num_in_box['Box_Type_ID']
                    dic_num_in_osc['A_Box_Type'] = dic_num_in_box['Box_Type']
                    dic_num_in_osc['A_ResPoint_Type_ID'] = dic_num_in_box['ResPoint_Type_ID']
                    dic_num_in_osc['A_Longitude'] = dic_num_in_box['Longitude']
                    dic_num_in_osc['A_Latitude'] = dic_num_in_box['Latitude']
                if dic_num_in_osc['Z_Box_Name'] == dic_num_in_box['Box_Name']:
                    dic_num_in_osc['Z_Box_Type_ID'] = dic_num_in_box['Box_Type_ID']
                    dic_num_in_osc['Z_Box_Type'] = dic_num_in_box['Box_Type']
                    dic_num_in_osc['Z_ResPoint_Type_ID'] = dic_num_in_box['ResPoint_Type_ID']
                    dic_num_in_osc['Z_Longitude'] = dic_num_in_box['Longitude']
                    dic_num_in_osc['Z_Latitude'] = dic_num_in_box['Latitude']

            dic_num_in_osc['Length'] = round(math.sqrt(((dic_num_in_osc['A_Longitude'] - dic_num_in_osc['Z_Longitude']) * Horizontal_Metre) ** 2 + ((dic_num_in_osc['A_Latitude'] - dic_num_in_osc['Z_Latitude'])* Vertical_Metre) ** 2))
            dic_num_in_osc['Business_Level'] = 8
            dic_num_in_osc['Life_Cycle'] = 8
            dic_num_in_osc['Owner_Type'] = 1
            dic_num_in_osc['Owner_Name'] = 0
            dic_num_in_osc['Field_Type'] = '市城区域'
            dic_num_in_osc['City_ID'] = List_Template_Selected[16]
            dic_num_in_osc['County_ID'] = List_Template_Selected[1]
            dic_num_in_osc['DQS_Project_ID'] = List_Template_Selected[3]
            dic_num_in_osc['DQS_ID'] = List_Template_Selected[5]
            dic_num_in_osc['DQS_County_ID'] = List_Template_Selected[7]
            dic_num_in_osc['DQS_Maintainer_ID'] = List_Template_Selected[9]
            dic_num_in_osc['Task_Name_ID'] = Task_Name_ID_List[0]

        '''生成List_OC_Data'''
        global List_OC_Data
        List_OC_Data = []
        for each_7013_line in List_7013: 
            if each_7013_line[10] == '二级':
                List_OC_Data.append(dict({
                    'Z_POS_Name': each_7013_line[11],
                    'Z_Box_Name': each_7013_line[3],
                    'A_POS_Name': each_7013_line[8],
                    'A_Port_Name': each_7013_line[9],
                    'City_ID': List_Template_Selected[16],
                    'County_ID': List_Template_Selected[1],
                    'Business_Name': Task_Name_ID_List[1],
                    'Project_Code': each_7013_line[4],
                    'Task_Name': each_7013_line[5],
                    'Task_Name_ID': Task_Name_ID_List[0],
                    'DQS_Project': List_Template_Selected[11],
                    'DQS_Project_ID': List_Template_Selected[3],
                    'DQS': List_Template_Selected[12],
                    'DQS_ID': List_Template_Selected[5],
                    'DQS_County': List_Template_Selected[13],
                    'DQS_County_ID': List_Template_Selected[7],
                    'DQS_Maintainer': List_Template_Selected[8],
                    'DQS_Maintainer_ID': List_Template_Selected[9],
                    'AEquType': 10002,
                    'ZEquType': 10002,
                    'SXBussType': '家客业务',
                    'AJoinName': '网元成端设备',
                    'ZJoinName': '网元成端设备',
                    'FiberCount': 1,
                    'BussType': 207, # 设备类型 207 PON/402 直联光路
                    'AppType': 1005, # 应用类型  1005 业务光路/1006 尾纤光路
                    'ServiceLevel': 101, # 101 家客接入/111 驻地网
                    }))
        for each_oc_data in List_OC_Data:
            for each_7013_line in List_7013:
                if each_oc_data['A_POS_Name'] == each_7013_line[11]: # 中文名称 each_7013_line[2] # 资管中文名称
                    each_oc_data['A_Box_Name'] = each_7013_line[3]
            if each_oc_data['A_Box_Name'] == each_oc_data['Z_Box_Name']:
                each_oc_data['BussType'] = 402
                each_oc_data['AppType'] = 1006
            for each_box_data in List_Box_Data:
                if each_oc_data['A_Box_Name'] == each_box_data['Box_Name']:
                    each_oc_data['A_Box_Type'] = each_box_data['Box_Type']
                    each_oc_data['A_Box_Type_ID'] = each_box_data['Box_Type_ID']
                    each_oc_data['A_ResPoint_Type_ID'] = each_box_data['ResPoint_Type_ID']
                if each_oc_data['Z_Box_Name'] == each_box_data['Box_Name']:
                    each_oc_data['Z_Box_Type'] = each_box_data['Box_Type']
                    each_oc_data['Z_Box_Type_ID'] = each_box_data['Box_Type_ID']
                    each_oc_data['Z_ResPoint_Type_ID'] = each_box_data['ResPoint_Type_ID']

def Generate_Topology():
    global CS_Topology
    CS_Topology = []
    Line_Num = 0
    Segment_Num = 0
    for cable_seg_num in range(len(List_CS_Data)):
        if cable_seg_num == 0:
            CS_Topology.append([List_CS_Data[cable_seg_num]['A_Box_Name']])
            CS_Topology[cable_seg_num].append(List_CS_Data[cable_seg_num]['Z_Box_Name'])
            Segment_Num += 1
            continue
        if List_CS_Data[cable_seg_num]['A_Box_Name'] == CS_Topology[Line_Num][Segment_Num]:
            CS_Topology[Line_Num].append(List_CS_Data[cable_seg_num]['Z_Box_Name'])
            Segment_Num += 1
            continue
        else:
            CS_Topology.append([List_CS_Data[cable_seg_num]['A_Box_Name']])
            Line_Num += 1
            Segment_Num = 1
            CS_Topology[Line_Num].append(List_CS_Data[cable_seg_num]['Z_Box_Name'])

def Generate_FS_Data():
    #filled by 0
    for box_info in List_Box_Data:
        box_info['1FS_Count'] = 0
        box_info['2FS_Count'] = 0
        box_info['DL_2FS_Count'] = 0
        box_info['ODM_Rows'] = 0
        box_info['Tray_Count'] = 0
    #1fs_count&2fs_count
    for row in List_7013:
        for box_info in List_Box_Data:
            if row[3] == box_info['Box_Name']:
                if row[10] == '一级':
                    box_info['1FS_Count'] += 1
                elif row[10] == '二级':
                    box_info['2FS_Count'] += 1
    #dl_2fs_count
    for cable_num in range(len(CS_Topology) - 1, -1, -1):
        for cable_seg_num in range(len(CS_Topology[cable_num]) - 1, -1, -1):
            if cable_seg_num == len(CS_Topology[cable_num]) - 1:#Tail
                for box_info in List_Box_Data: 
                    if CS_Topology[cable_num][cable_seg_num] == box_info['Box_Name']:
                        box_info['DL_2FS_Count'] = 0
            elif cable_seg_num == 0:#Head
                for box_info in List_Box_Data:
                    if CS_Topology[cable_num][cable_seg_num] == box_info['Box_Name']:
                        box_info['DL_2FS_Count'] = 0
            else:#Middle
                for box_info in List_Box_Data:
                    if CS_Topology[cable_num][cable_seg_num + 1] == box_info['Box_Name']:
                        DL_2FS_Count_temp = box_info['2FS_Count'] + box_info['DL_2FS_Count']
                for box_info in List_Box_Data:
                    if CS_Topology[cable_num][cable_seg_num] == box_info['Box_Name']:
                        box_info['DL_2FS_Count'] = DL_2FS_Count_temp
    #dl_2fs_count 1fs_box_correction
    for box_info in List_Box_Data:
        if box_info['1FS_Count'] != 0:
            DL_2FS_Count_temp = []
            for cable_num in CS_Topology:
                if box_info['Box_Name'] == cable_num[0]:
                    for  box_info_2 in List_Box_Data:
                        if cable_num[1] == box_info_2['Box_Name']:
                            DL_2FS_Count_temp.append(str(box_info_2['DL_2FS_Count'] + box_info_2['2FS_Count']))
                            box_info['DL_2FS_Count'] = DL_2FS_Count_temp
    #ODM_Rows$Tray_Count
    for box_info in List_Box_Data:
        if box_info['1FS_Count'] == 0:
            for cable_seg_data in List_CS_Data:
                if box_info['Box_Name'] == cable_seg_data['Z_Box_Name']:
                    box_info['ODM_Rows'] = box_info['Tray_Count'] = math.ceil(cable_seg_data['Width'] / 12)
        elif box_info['1FS_Count'] != 0:
            for cable_seg_data in List_CS_Data:
                if box_info['Box_Name'] == cable_seg_data['A_Box_Name']:
                    box_info['ODM_Rows'] = box_info['Tray_Count'] = math.ceil(cable_seg_data['Width'] / 12) + box_info['ODM_Rows']

    # 根据照片修改数据
    for box_info in List_Box_Data:
        for modify_tray in List_Modify_For_Photo:
            if box_info['Box_Name'] == modify_tray[0]:
                 box_info['ODM_Rows'] = box_info['Tray_Count'] = modify_tray[1]

def Generate_Termination_and_Direct_Melt_Data():

    for box_info in List_Box_Data:

        if box_info['Box_Type'] == 'guangfenxianxiang':
            box_info['Box_Type_Short'] = 'gfxx'
        elif box_info['Box_Type'] == 'guangjiaojiexiang':
            box_info['Box_Type_Short'] = 'gjjx'

        # 基本上架直熔数据
        if box_info['1FS_Count'] == 0:
            for cable_seg_data in List_CS_Data:
                if box_info['Box_Name'] == cable_seg_data['Z_Box_Name']:
                    if cable_seg_data['Width'] >= (box_info['2FS_Count'] * 3 + box_info['DL_2FS_Count'] * 3):
                        box_info['BackUp_Fiber_Count'] = 2
                    elif (cable_seg_data['Width'] < (box_info['2FS_Count'] * 3 + box_info['DL_2FS_Count'] * 3)) and (cable_seg_data['Width'] >= (box_info['2FS_Count'] * 2 + box_info['DL_2FS_Count'] * 2)):
                        box_info['BackUp_Fiber_Count'] = 1
                    else:
                        box_info['BackUp_Fiber_Count'] = 0
            box_info['Termination_Start'] = 1
            box_info['Termination_Count'] = box_info['2FS_Count'] * (box_info['BackUp_Fiber_Count'] + 1)
            box_info['Direct_Melt_Start'] = box_info['2FS_Count'] * (box_info['BackUp_Fiber_Count'] + 1) + 1
            box_info['Direct_Melt_Count'] = box_info['DL_2FS_Count'] * (box_info['BackUp_Fiber_Count'] + 1)
            if box_info['Direct_Melt_Count'] == 0: #尾箱没有直熔数据
                box_info['Direct_Melt_Start'] = 0

        elif box_info['1FS_Count'] != 0:

            List_Width = []
            for cable_seg_data in List_CS_Data:
                if box_info['Box_Name'] == cable_seg_data['A_Box_Name']:
                    List_Width.append(cable_seg_data['Width'])

            box_info['BackUp_Fiber_Count'] = []
            box_info['Termination_Start'] = []
            box_info['Termination_Count'] = []
            box_info['Direct_Melt_Start'] = []
            box_info['Direct_Melt_Count'] = []

            for each_DL_2FS, each_Width in zip(box_info['DL_2FS_Count'], List_Width):
                if each_Width >= (box_info['2FS_Count'] * 3 + int(each_DL_2FS) * 3):
                    box_info['BackUp_Fiber_Count'].append(2)
                elif (each_Width < (box_info['2FS_Count'] * 3 + int(each_DL_2FS) * 3)) and (each_Width >= (box_info['2FS_Count'] * 2 + int(each_DL_2FS) * 2)):
                    box_info['BackUp_Fiber_Count'].append(1)
                else:
                    box_info['BackUp_Fiber_Count'].append(0)
            
            for each_BackUp_Fiber_Count, each_DL_2FS in zip(box_info['BackUp_Fiber_Count'], box_info['DL_2FS_Count']):
                box_info['Termination_Start'].append(1)
                box_info['Termination_Count'].append(int(each_DL_2FS) * (int(each_BackUp_Fiber_Count) + 1))
                box_info['Direct_Melt_Start'] = '0'
                box_info['Direct_Melt_Count'] = '0'
            
            box_info['BackUp_Fiber_Count'] = [str(i) for i in box_info['BackUp_Fiber_Count']]
            box_info['Termination_Start'] = [str(i) for i in box_info['Termination_Start']]
            box_info['Termination_Count'] = [str(i) for i in box_info['Termination_Count']]
            box_info['Direct_Melt_Start'] = [str(i) for i in box_info['Direct_Melt_Start']]
            box_info['Direct_Melt_Count'] = [str(i) for i in box_info['Direct_Melt_Count']]

        # Termination_Sequence处理

        box_info['Termination_Sequence'] = []
        if isinstance(box_info['Termination_Count'], int): #单条下行光缆
            for terminal_sequence in range(int(box_info['Termination_Count'])):
                box_info['Termination_Sequence'].append(terminal_sequence)

            # 处理纤芯
            List_CS_Fiber_IDs =[]
            box_info['Termination_Fiber_IDs'] = []
            for cable_seg_data in List_CS_Data:
                if box_info['Box_Name'] == cable_seg_data['Z_Box_Name']:
                    List_CS_Fiber_IDs = cable_seg_data['CS_Fiber_IDs']
            box_info['Termination_Fiber_IDs'] = List_CS_Fiber_IDs[0:int(box_info['Termination_Count'])]

        elif isinstance(box_info['Termination_Count'], list): #多条下行光缆

            Int_Termination_Sequence_Pointer = 0 # 利用指针生成上架数据
            box_info['Termination_Sequence'] = []
            for sub_count in box_info['Termination_Count']:
                for terminal_sequence in range(int(sub_count)):

                    # 操作指针
                    if Int_Termination_Sequence_Pointer % 12 != 0:
                        for i in range(11):
                            i += 0
                            Int_Termination_Sequence_Pointer += 1
                            if Int_Termination_Sequence_Pointer % 12 == 0:
                                break

                    # 写入上架
                    box_info['Termination_Sequence'].append(terminal_sequence + Int_Termination_Sequence_Pointer)
                Int_Termination_Sequence_Pointer += int(sub_count)

            # 处理纤芯
            List_CS_Fiber_IDs =[]
            box_info['Termination_Fiber_IDs'] = []
            for cable_seg_data in List_CS_Data:
                if box_info['Box_Name'] == cable_seg_data['A_Box_Name']:
                    List_CS_Fiber_IDs.append(cable_seg_data['CS_Fiber_IDs'])
            for sub_count, sub_list in zip(box_info['Termination_Count'], List_CS_Fiber_IDs):
                List_Fiber_IDs = sub_list[0:int(sub_count)]
                for fiber_id in List_Fiber_IDs:
                    box_info['Termination_Fiber_IDs'].append(fiber_id)

            # 端子纤芯对应(不看照片的话,这里的数据可以直接上架)
            List_Terminal_ID_and_Fiber_ID_All = []
            for fiber_sequence in range(len(box_info['Termination_Sequence'])):
                List_Terminal_ID_and_Fiber_ID_All.append(list([box_info['Termination_Sequence'][fiber_sequence],box_info['Terminal_IDs'][int(box_info['Termination_Sequence'][fiber_sequence])],fiber_sequence,box_info['Termination_Fiber_IDs'][fiber_sequence]]))

            # 分离占用与备芯
            # slice数据
            List_Terminal_ID_and_Fiber_ID_Occupied = []
            List_Terminal_ID_and_Fiber_ID_Free = copy.deepcopy(List_Terminal_ID_and_Fiber_ID_All)
            List_Slice_Data = []
            for sub_count, sub_backup_fiber_count in zip(box_info['Termination_Count'], box_info['BackUp_Fiber_Count']):
                List_Slice_Data.append([(int(sub_backup_fiber_count) + 1), int(sub_count)])
            Int_Terminaion_Slice_Pointer = 0
            for sub_list in List_Slice_Data:
                for num in range(int(0 + Int_Terminaion_Slice_Pointer),int(sub_list[1] + Int_Terminaion_Slice_Pointer),int(sub_list[0])):
                    List_Terminal_ID_and_Fiber_ID_Occupied.append(List_Terminal_ID_and_Fiber_ID_All[num])
                Int_Terminaion_Slice_Pointer += sub_list[1]
            for sub_list_all in List_Terminal_ID_and_Fiber_ID_Free:
                for sub_list_occupied in List_Terminal_ID_and_Fiber_ID_Occupied:
                    if sub_list_all == sub_list_occupied:
                        List_Terminal_ID_and_Fiber_ID_Free.remove(sub_list_all)
            # 注意!注意!注意!
            box_info['Direct_Melt_Count'] = List_Terminal_ID_and_Fiber_ID_Occupied
            box_info['Direct_Melt_Start'] = List_Terminal_ID_and_Fiber_ID_Free
            # box_info['Direct_Melt_Count'] = List_Terminal_ID_and_Fiber_ID_Occupied + List_Terminal_ID_and_Fiber_ID_Free

           # 根据照片修改数据
            if len(List_Modify_For_Photo) != 0:
                List_Terminal_ID_and_Fiber_ID_Occupied_Modify = []
                for modify_data in List_Modify_For_Photo:
                    if modify_data[0] == box_info['Box_Name']:
                        List_Modidy_Termination = [(int(i) - 1) for i in modify_data[2].split(',')]

                        # 自定义上架写入Terminal_ID
                        for terminal_id_sequence in List_Modidy_Termination:
                            List_Terminal_ID_and_Fiber_ID_Occupied_Modify.append([terminal_id_sequence,box_info['Terminal_IDs'][terminal_id_sequence]])
                        for sub_list, sub_list_modify in zip(List_Terminal_ID_and_Fiber_ID_Occupied, List_Terminal_ID_and_Fiber_ID_Occupied_Modify):
                            sub_list_modify.append(sub_list[2])
                            sub_list_modify.append(sub_list[3])

                        # 生成备芯数据,为了思路清晰,分步处理

                        # 分离自定占用的端子和纤芯
                        List_Terminal_ID_Occupied_Modify = []
                        List_Fiber_ID_Occupied_Modify = []
                        for sequence in List_Terminal_ID_and_Fiber_ID_Occupied_Modify:
                            List_Terminal_ID_Occupied_Modify.append([sequence[0],sequence[1]])
                            List_Fiber_ID_Occupied_Modify.append([sequence[2],sequence[3]])

                        # 分离原始全部的端子和纤芯
                        List_Terminal_ID_All = []
                        List_Fiber_ID_All = []
                        for sequence in List_Terminal_ID_and_Fiber_ID_All:
                            List_Terminal_ID_All.append([sequence[0],sequence[1]])
                            List_Fiber_ID_All.append([sequence[2],sequence[3]])

                        # 生成备芯数据
                        List_Terminal_ID_Free_Modify = copy.deepcopy(List_Terminal_ID_All)
                        List_Fiber_ID_Free_Modify = copy.deepcopy(List_Fiber_ID_All)

                        # 端子
                        for sequence_occupied in List_Terminal_ID_Occupied_Modify:
                            for sequence_free in List_Terminal_ID_Free_Modify:
                                if sequence_free[1] == sequence_occupied[1]:
                                    List_Terminal_ID_Free_Modify.remove(sequence_free)

                        # 纤芯
                        for sequence_occupied in List_Fiber_ID_Free_Modify:
                            for sequence_free in List_Fiber_ID_Occupied_Modify:
                                if sequence_free[1] == sequence_occupied[1]:
                                    List_Fiber_ID_Free_Modify.remove(sequence_free)

                        # 合并端子和纤芯
                        List_Terminal_ID_and_Fiber_ID_Free_Modify = []
                        for terminal_id_sequence, fiber_id_sequence in zip(List_Terminal_ID_Free_Modify, List_Fiber_ID_Free_Modify):
                            List_Terminal_ID_and_Fiber_ID_Free_Modify.append([terminal_id_sequence[0], terminal_id_sequence[1], fiber_id_sequence[0], fiber_id_sequence[1]])
                        # 注意!注意!注意!
                        box_info['Direct_Melt_Count'] = List_Terminal_ID_and_Fiber_ID_Occupied_Modify
                        box_info['Direct_Melt_Start'] = List_Terminal_ID_and_Fiber_ID_Free_Modify
                        # box_info['Direct_Melt_Count'] = List_Terminal_ID_and_Fiber_ID_Occupied_Modify + List_Terminal_ID_and_Fiber_ID_Free_Modify

def Generate_OC_POS_Data_and_OC_Name():
    List_OC_Name = []
    for each_oc_data in List_OC_Data:
        for each_box_data in List_Box_Data:
            if each_oc_data['A_Box_Name'] == each_box_data['Box_Name']:
                for keykey, valuevalue in each_box_data['POS_IDs'].items():
                    if each_oc_data['A_POS_Name'] == keykey:
                        each_oc_data['A_POS_ID'] = valuevalue
            if each_oc_data['Z_Box_Name'] == each_box_data['Box_Name']:
                for keykey, valuevalue in each_box_data['POS_IDs'].items():
                    if each_oc_data['Z_POS_Name'] == keykey:
                        each_oc_data['Z_POS_ID'] = valuevalue

        each_oc_data['OC_Name'] = each_oc_data['A_Box_Name'] + '资源点' + '-' + each_oc_data['Z_Box_Name'] + '资源点' + 'F0001'
        List_OC_Name.append(each_oc_data['OC_Name'])

    Counter_OC_Name = collections.Counter()
    for oc_name in List_OC_Name:
        Counter_OC_Name[oc_name] += 1

    for keykey, valuevalue in Counter_OC_Name.items():
        for increase_num in range(valuevalue):
            for each_oc_data in List_OC_Data:
                if each_oc_data['OC_Name'] == keykey:
                    each_oc_data['OC_Name'] = each_oc_data['OC_Name'][:(len(each_oc_data['OC_Name']) - 4)] + '{:04d}'.format(int(each_oc_data['OC_Name'][(len(each_oc_data['OC_Name']) - 4):]) + int(increase_num))
                    break


def Query_Project_Code_ID():
    URL_Query_Project_Code_ID = 'http://10.209.199.74:8120/igisserver_osl/rest/datatrans/expall?model=guangfenxianxiang&fname=PROJECTCODE&p1='+List_7013[1][4]
    Response_Body = requests.get(URL_Query_Project_Code_ID)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Response_Key = Response_Body.xpath("//@name")
    List_Response_Value = Response_Body.xpath("//@value")
    Dic_Response = dict(zip(List_Response_Key,List_Response_Value))
    Dic_Project_Code_ID = copy.deepcopy(Dic_Response)
    for box_num in List_Box_Data:
        box_num['Project_Code_ID'] = Dic_Project_Code_ID[List_7013[1][4]]
    for ocs_num in List_CS_Data:
        ocs_num['Project_Code_ID'] = Dic_Project_Code_ID[List_7013[1][4]]
    for oc_data in List_OC_Data:
        oc_data['Project_Code_ID'] = Dic_Project_Code_ID[List_7013[1][4]]

def Query_Box_ID_ResPoint_ID_Alias(Para_List_Box_Data):
    URL_Query_Box = 'http://10.209.199.74:8120/igisserver_osl/rest/generalSaveOrGet/generalGet'
    Form_Info = '<request><query mc="'+Para_List_Box_Data['Box_Type']+'" ids="" where="1=1 AND ZH_LABEL LIKE \'%'+Para_List_Box_Data['Box_Name']+'%\'" returnfields="INT_ID,ZH_LABEL,STRUCTURE_ID,ALIAS"/></request>'
    Form_Info_Encoded = 'xml='+urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {"Host": "10.209.199.74:8120","Content-Type": "application/x-www-form-urlencoded","Content-Length": Request_Lenth}
    Response_Body = requests.post(URL_Query_Box, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Response_Key = Response_Body.xpath("//fv/@k")
    List_Response_Value = Response_Body.xpath("//fv/@v")
    List_Response_Value_tv = Response_Body.xpath("//fv/@tv")
    Dic_Response = dict(zip(List_Response_Key,List_Response_Value))
    for box_num in range(len(List_Box_Data)):
        if Dic_Response['ZH_LABEL'] == List_Box_Data[box_num]['Box_Name']:
            List_Box_Data[box_num]['Box_ID'] = Dic_Response['INT_ID']
            List_Box_Data[box_num]['ResPoint_ID'] = Dic_Response['STRUCTURE_ID']
            List_Box_Data[box_num]['Alias'] = Dic_Response['ALIAS']
            List_Box_Data[box_num]['ResPoint_Name'] = List_Response_Value_tv[0]

    if P0_Data_Check or P2_Generate_Support_Segment or P3_Generate_Cable_Segment or P4_Cable_Lay or P5_Generate_ODM or P6_Generate_Tray or P7_Termination or P8_Direct_Melt or P9_Generate_Optical_Circuit or P10_Transmission_Design or P11_Termination or P12_Update_1_Fix_OCS:
        for ocs_num in List_CS_Data:
            for box_num in List_Box_Data:
                if ocs_num['A_Box_Name'] == box_num['Box_Name']:
                    ocs_num['A_Box_ID'] = box_num['Box_ID']
                    ocs_num['A_ResPoint_ID'] = box_num['ResPoint_ID']
            for box_num in List_Box_Data:
                if ocs_num['Z_Box_Name'] == box_num['Box_Name']:
                    ocs_num['Z_Box_ID'] = box_num['Box_ID']
                    ocs_num['Z_ResPoint_ID'] = box_num['ResPoint_ID']
        for each_oc_data in List_OC_Data:
            for each_box_data in List_Box_Data:
                if each_oc_data['A_Box_Name'] == each_box_data['Box_Name']:
                    each_oc_data['A_Box_ID'] = each_box_data['Box_ID']
                    each_oc_data['A_ResPoint_ID'] = each_box_data['ResPoint_ID']
                    each_oc_data['A_ResPoint_Name'] = each_box_data['ResPoint_Name']
                if each_oc_data['Z_Box_Name'] == each_box_data['Box_Name']:
                    each_oc_data['Z_Box_ID'] = each_box_data['Box_ID']
                    each_oc_data['Z_ResPoint_ID'] = each_box_data['ResPoint_ID']
                    each_oc_data['Z_ResPoint_Name'] = each_box_data['ResPoint_Name']

def Query_Support_Sys_and_Cable_Sys():
    Task_Name_ID_List = List_7013[1][5].split('-')
    Support_Sys_Name = List_7013[1][0]+List_7013[1][1]+Task_Name_ID_List[1]+'驻地网引上'
    Cable_Sys_Name = List_7013[1][0]+List_7013[1][1]+Task_Name_ID_List[1]+'驻地网架空光缆'

    #Support_System
    URL_Query_SS_ID = 'http://10.209.199.74:8120/igisserver_osl/rest/generalSaveOrGet/generalGetPage'
    Form_Info = '<request><query mc="xitong" where="1=1 AND LINESEG_TYPE=\'9108\' AND ZH_LABEL LIKE \'%'+Support_Sys_Name+'%\'" returnfields="INT_ID,ZH_LABEL"/></request>'
    Form_Info_Encoded = 'pageNum=1&'+'xml='+urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Query_SS_ID, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_SS_Count = Response_Body.xpath('//@counts')
    if List_SS_Count[0] == '0':
        #Generate Support_Sys
        URL_Add_Support_Sys = 'http://10.209.199.74:8120/igisserver_osl/rest/generalSaveOrGet/generalSave'
        Form_Info = '<response mode="add"><mc type="xitong"><mo group="1"><fv k="SYSTEM_LEVEL" v="8"/><fv k="CITY_ID" v="'+str(List_CS_Data[0]['City_ID'])+'"/><fv k="STATUS" v="8"/><fv k="COUNTY_ID" v="'+str(List_CS_Data[0]['County_ID'])+'"/><fv k="LINESEG_TYPE" v="9108"/><fv k="ZH_LABEL" v="'+Support_Sys_Name+'"/><fv k="INT_ID" v="new-27991311"/></mo></mc></response>'
        Form_Info_Encoded = 'xml='+urllib.parse.quote_plus(Form_Info)
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Add_Support_Sys, data=Form_Info_Encoded, headers=Request_Header)
        Response_Body = bytes(Response_Body.text, encoding="utf-8")
        Response_Body = lxml.etree.HTML(Response_Body)
        Support_Sys_ID = Response_Body.xpath('//@newid')
        Support_Sys_ID = Support_Sys_ID[0]
        for ocs_num in List_CS_Data:
            ocs_num['Support_Sys_ID'] = Support_Sys_ID
        print('引上系统ID-{}'.format(Support_Sys_ID))

    elif List_SS_Count[0] == '1':
        #Get Support_Sys_ID
        Support_Sys_ID = Response_Body.xpath('//@int_id')
        Support_Sys_ID = Support_Sys_ID[0]
        for ocs_num in List_CS_Data:
            ocs_num['Support_Sys_ID'] = Support_Sys_ID
        print('引上系统ID-{}'.format(Support_Sys_ID))

    #Cable_System
    URL_Query_CS_ID = 'http://10.209.199.74:8120/igisserver_osl/rest/generalSaveOrGet/generalGetPage'
    Form_Info = '<request><query mc="guanglanxitong" where="1=1 AND ZH_LABEL LIKE \'%'+Cable_Sys_Name+'%\'" returnfields="INT_ID,ZH_LABEL"/></request>'
    Form_Info_Encoded = 'pageNum=1&'+'xml='+urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Query_CS_ID, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_CS_Count = Response_Body.xpath('//@counts')
    if List_CS_Count[0] == '0':
        #Generate Cable_Sys
        URL_Add_Cable_Sys = 'http://10.209.199.74:8120/igisserver_osl/rest/generalSaveOrGet/generalSave'
        Form_Info = '<response mode="add"><mc type="xitong"><mo group="1"><fv k="SYSTEM_LEVEL" v="8"/><fv k="CITY_ID" v="'+str(List_CS_Data[0]['City_ID'])+'"/><fv k="STATUS" v="8"/><fv k="COUNTY_ID" v="'+str(List_CS_Data[0]['County_ID'])+'"/><fv k="LINESEG_TYPE" v="9108"/><fv k="ZH_LABEL" v="'+Support_Sys_Name+'"/><fv k="INT_ID" v="new-27991311"/></mo></mc></response>'
        Form_Info = '<response mode="add"><mc type="guanglanxitong"><mo group="1"><fv k="SYSTEM_LEVEL" v="8"/><fv k="CITY_ID" v="'+str(List_CS_Data[0]['City_ID'])+'"/><fv k="STATUS" v="8"/><fv k="COUNTY_ID" v="'+str(List_CS_Data[0]['County_ID'])+'"/><fv k="ZH_LABEL" v="'+Cable_Sys_Name+'"/><fv k="INT_ID" v="new-27991311"/><fv k="OPT_TYPE" v="GYTA-12"/></mo></mc></response>'
        Form_Info_Encoded = 'xml='+urllib.parse.quote_plus(Form_Info)
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Add_Cable_Sys, data=Form_Info_Encoded, headers=Request_Header)
        Response_Body = bytes(Response_Body.text, encoding="utf-8")
        Response_Body = lxml.etree.HTML(Response_Body)
        Cable_Sys_ID = Response_Body.xpath('//@newid')
        Cable_Sys_ID = Cable_Sys_ID[0]
        for ocs_num in List_CS_Data:
            ocs_num['Cable_Sys_ID'] = Cable_Sys_ID
        print('光缆系统ID-{}'.format(Cable_Sys_ID))

    elif List_CS_Count[0] == '1':
        #Get Cable_Sys_ID
        Cable_Sys_ID = Response_Body.xpath('//@int_id')
        Cable_Sys_ID = Cable_Sys_ID[0]
        for ocs_num in List_CS_Data:
            ocs_num['Cable_Sys_ID'] = Cable_Sys_ID
        print('光缆系统ID-{}'.format(Cable_Sys_ID))

def Query_Run_Certification():

    def add_to_16(Para_Password):
        while len(Para_Password) % 16 != 0:
            Para_Password += '\x00'
        return str.encode(Para_Password)

    Encrypt_key='sxportaljiamikey'
    Encrypt_key_byte = Encrypt_key.encode()
    Encrypt_iv ='sxportaljiamiwyl'
    Encrypt_iv_byte = Encrypt_iv.encode()
    print(Username_Run)
    print(Password_Run)
    Login_Password_byte = Password_Run.encode()
    Login_Password_byte = add_to_16(Password_Run)
    Cipher = Crypto.Cipher.AES.new(Encrypt_key_byte, Crypto.Cipher.AES.MODE_CBC, Encrypt_iv_byte)
    Encrypted_byte = Cipher.encrypt(Login_Password_byte)
    Encrypted_byte = base64.b64encode(Encrypted_byte)
    Encrypt_Password = Encrypted_byte.decode()

    session = requests.session()
    Response_Body = session.post('http://portal.sx.cmcc/pkmslogin.form?uid=' + Username_Run,data= {'login-form-type':'pwd','username':Username_Run,'password':Encrypt_Password})
    Response_Body = session.post('http://portal.sx.cmcc/sxmcc_wcm/middelwebpage/encryptportallogin/encryptlogin.jsp?appurl=http://10.209.199.72:7112/irms/sso.action')
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_userdt_ipaddress = Response_Body.xpath('//@value')
    Response_Body = requests.post('http://10.209.199.72:7112/irms/sso.action',data={'userdt': List_userdt_ipaddress[0], 'ipAddress': List_userdt_ipaddress[1]})
    cookies = requests.utils.dict_from_cookiejar(Response_Body.cookies)
    global Jsessionirms_v_Run, route_v_Run
    Jsessionirms_v_Run = cookies['JSESSIONIRMS']
    route_v_Run = cookies['route']
    print('JSESSIONIRMS_Run: ' + Jsessionirms_v_Run)
    print('route_Run: ' + route_v_Run)

def Query_Create_Certification():

    def add_to_16(Para_Password):
        while len(Para_Password) % 16 != 0:
            Para_Password += '\x00'
        return str.encode(Para_Password)

    Encrypt_key='sxportaljiamikey'
    Encrypt_key_byte = Encrypt_key.encode()
    Encrypt_iv ='sxportaljiamiwyl'
    Encrypt_iv_byte = Encrypt_iv.encode()
    print(Username_Create)
    print(Password_Create)
    Login_Password_byte = Password_Create.encode()
    Login_Password_byte = add_to_16(Password_Create)
    Cipher = Crypto.Cipher.AES.new(Encrypt_key_byte, Crypto.Cipher.AES.MODE_CBC, Encrypt_iv_byte)
    Encrypted_byte = Cipher.encrypt(Login_Password_byte)
    Encrypted_byte = base64.b64encode(Encrypted_byte)
    Encrypt_Password = Encrypted_byte.decode()

    session = requests.session()
    Response_Body = session.post('http://portal.sx.cmcc/pkmslogin.form?uid=' + Username_Create,data= {'login-form-type':'pwd','username':Username_Create,'password':Encrypt_Password})
    Response_Body = session.post('http://portal.sx.cmcc/sxmcc_wcm/middelwebpage/encryptportallogin/encryptlogin.jsp?appurl=http://10.209.199.72:7112/irms/sso.action')
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_userdt_ipaddress = Response_Body.xpath('//@value')
    Response_Body = requests.post('http://10.209.199.72:7112/irms/sso.action',data={'userdt': List_userdt_ipaddress[0], 'ipAddress': List_userdt_ipaddress[1]})
    cookies = requests.utils.dict_from_cookiejar(Response_Body.cookies)
    global Jsessionirms_v_Create, route_v_Create
    Jsessionirms_v_Create = cookies['JSESSIONIRMS']
    route_v_Create = cookies['route']
    print('JSESSIONIRMS_Create: ' + Jsessionirms_v_Create)
    print('route_Create: ' + route_v_Create)

def Query_Support_Seg_ID_Cable_Seg_ID(Para_List_CS_Data):
    List_CS_Support_Seg_Name_ID_Cable_Name_ID = {}
    URL_Query_Support_Seg_ID_Cable_Seg_ID = 'http://10.209.199.74:8120/igisserver_osl/rest/generalSaveOrGet/generalGet'

    Form_Info = '<request><query mc="yinshangduan" where="1=1 AND ZH_LABEL LIKE \'%'+str(Para_List_CS_Data['A_Box_Name'])+'资源点-'+str(Para_List_CS_Data['Z_Box_Name'])+'资源点'+'%\'" returnfields="INT_ID,ZH_LABEL"/></request>'
    Form_Info_Encoded = 'xml='+urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {"Host": "10.209.199.74:8120","Content-Type": "application/x-www-form-urlencoded","Content-Length": Request_Lenth}
    Response_Body = requests.post(URL_Query_Support_Seg_ID_Cable_Seg_ID, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Response_Value = Response_Body.xpath("//fv/@v")
    List_CS_Support_Seg_Name_ID_Cable_Name_ID['Support_Seg_Name'] = List_Response_Value[1]
    List_CS_Support_Seg_Name_ID_Cable_Name_ID['Support_Seg_ID'] = List_Response_Value[0]

    Form_Info = '<request><query mc="guanglanduan" where="1=1 AND ZH_LABEL LIKE \'%'+str(Para_List_CS_Data['A_Box_Name'])+'资源点-'+str(Para_List_CS_Data['Z_Box_Name'])+'资源点'+'%\'" returnfields="INT_ID,ZH_LABEL"/></request>'
    Form_Info_Encoded = 'xml='+urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {"Host": "10.209.199.74:8120","Content-Type": "application/x-www-form-urlencoded","Content-Length": Request_Lenth}
    Response_Body = requests.post(URL_Query_Support_Seg_ID_Cable_Seg_ID, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Response_Value = Response_Body.xpath("//fv/@v")
    List_CS_Support_Seg_Name_ID_Cable_Name_ID['Cable_Seg_Name'] = List_Response_Value[1]
    List_CS_Support_Seg_Name_ID_Cable_Name_ID['Cable_Seg_ID'] = List_Response_Value[0]
    for cable_seg in List_CS_Data:
        if (Para_List_CS_Data['A_Box_Name'] == cable_seg['A_Box_Name']) and (Para_List_CS_Data['Z_Box_Name'] == cable_seg['Z_Box_Name']):
            cable_seg['Support_Seg_Name'] = List_CS_Support_Seg_Name_ID_Cable_Name_ID['Support_Seg_Name']
            cable_seg['Support_Seg_ID'] = List_CS_Support_Seg_Name_ID_Cable_Name_ID['Support_Seg_ID']
            cable_seg['Cable_Seg_Name'] = List_CS_Support_Seg_Name_ID_Cable_Name_ID['Cable_Seg_Name']
            cable_seg['Cable_Seg_ID'] = List_CS_Support_Seg_Name_ID_Cable_Name_ID['Cable_Seg_ID']

def Query_ODM_ID_and_Terminarl_IDs(Para_List_Box_Data):
    URL_Query_ODM_ID_and_Terminarl_IDs = 'http://10.209.199.74:8120/nxapp/room/queryShelfAndModuleData.ilf'
    if Para_List_Box_Data['Box_Type'] == 'guangfenxianxiang':
        Para_List_Box_Data['Box_Type_Short'] = 'gfxx'
    elif Para_List_Box_Data['Box_Type'] == 'guangjiaojiexiang':
        Para_List_Box_Data['Box_Type_Short'] = 'gjjx'
    Form_Info = '<params><param key="type" value="'+Para_List_Box_Data['Box_Type_Short']+'"/><param key="rack_id" value="'+Para_List_Box_Data['Box_ID']+'"/></params>'
    Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
    Form_Info_Encoded = "params=" + urllib.parse.quote_plus(Form_Info) + "&model=odm&" +  "lifeparams=" + urllib.parse.quote_plus(Form_Info_Tail)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Query_ODM_ID_and_Terminarl_IDs, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = urllib.parse.unquote(bytes(Response_Body.text, encoding="utf-8"))
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Terminal_IDs = Response_Body.xpath('//@id')
    List_Terminal_IDs.pop(0)
    Para_List_Box_Data['ODM_ID'] = List_Terminal_IDs.pop(0)
    Para_List_Box_Data['Terminal_IDs'] = List_Terminal_IDs

def Query_CS_Fiber_IDs(Para_List_CS_Data):
    URL_Query_CS_Fiber_IDs = 'http://10.209.199.74:8120/igisserver_osl/rest/EquipEditModule1/getEquipModuleTerminals'
    Form_Info = '<params><param key="equ_type" value="fiberseg"/><param key="equ_id" value="'+Para_List_CS_Data['Cable_Seg_ID']+'"/></params>'
    Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
    Form_Info_Encoded = "params=" + urllib.parse.quote_plus(Form_Info) + "&lifeparams=" + urllib.parse.quote_plus(Form_Info_Tail)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Query_CS_Fiber_IDs, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_CS_Fiber_IDs = Response_Body.xpath('//@id')
    List_CS_Fiber_IDs.pop(0)
    Para_List_CS_Data['CS_Fiber_IDs'] = List_CS_Fiber_IDs

def Query_POS_ID(Para_List_Box_Data):
    URL_Query_POS_ID = 'http://10.209.199.72:7112/irms/opticOpenApplyAction!queryEqu.ilf'
    Form_Info_Encoded = 'limit=30' + '&equType=POS&countyId=' + str(Para_List_Box_Data['County_ID']) + '&siteId=' + str(Para_List_Box_Data['ResPoint_ID']) + '&sitetype=' + str(Para_List_Box_Data['ResPoint_Type_ID']) + '&sitename=' + urllib.parse.quote_plus(Para_List_Box_Data['ResPoint_Name']) + '&cityId=' + str(Para_List_Box_Data['City_ID'])
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Query_POS_ID, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
    Response_Body = Response_Body.text
    Response_Body = Response_Body.replace('success:true','"success":true')
    Response_Body = Response_Body.replace('totalProperty','"totalProperty"')
    Response_Body = Response_Body.replace('data','"data"')
    Response_Body = Response_Body.replace('int_id','"int_id"')
    Response_Body = Response_Body.replace('endEquipType','"endEquipType"')
    Response_Body = Response_Body.replace('trans_site_type','"trans_site_type"')
    Response_Body = Response_Body.replace('trans_site_id','"trans_site_id"')
    Response_Body = Response_Body.replace('trans_site_name','"trans_site_name"')
    Response_Body = Response_Body.replace('zh_label','"zh_label"')
    Response_Body = Response_Body.replace('\'','\"')
    Response_Body = json.loads(Response_Body)
    List_POS_Name = []
    List_POS_ID = []
    for each_POS_data in Response_Body['data']:
        List_POS_Name.append(each_POS_data['zh_label'])
        List_POS_ID.append(each_POS_data['int_id'])
    Para_List_Box_Data['POS_IDs'] = dict(zip(List_POS_Name, List_POS_ID))

def Query_POS_Port_IDs(Para_List_Box_Data):
    Para_List_Box_Data['POS_Port_IDs'] = []
    for valuevalue in Para_List_Box_Data['POS_IDs'].values():
        URL_Query_POS_Port_IDs = 'http://10.209.199.74:8120/igisserver_osl/rest/EquipEditModule1/getEquipModuleTerminals'
        Form_Info = '<params><param key="equ_type" value="pos"/><param key="equ_id" value="'+str(valuevalue)+'"/></params>'
        Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
        Form_Info_Encoded = "params=" + urllib.parse.quote_plus(Form_Info) + "&lifeparams=" + urllib.parse.quote_plus(Form_Info_Tail)
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Query_POS_Port_IDs, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
        Response_Body = bytes(Response_Body.text, encoding="utf-8")
        Response_Body = lxml.etree.HTML(Response_Body)
        List_POS_Port_IDs = Response_Body.xpath('//@id')
        List_POS_Port_IDs.pop(0)
        List_POS_Port_Names = Response_Body.xpath('//@name')
        Para_List_Box_Data['POS_Port_IDs'].append(List_POS_Port_IDs)
        
        #写入A_Port_ID,Z_Port_Name,Z_Port_ID
        Dic_Port_Name_and_Port_ID = dict(zip(List_POS_Port_Names, List_POS_Port_IDs))
        RE_Pattern = re.compile(r'^.*分光器\d{3}|^.*分光器')
        for each_oc_data in List_OC_Data:
            for each_name,each_id in Dic_Port_Name_and_Port_ID.items():
                if each_oc_data['A_Port_Name'] == each_name:
                    each_oc_data['A_Port_ID'] = each_id

                Temp_Data = RE_Pattern.findall(each_oc_data['Z_POS_Name'])
                Temp_Data = Temp_Data[0] + '-0-001'
                if Temp_Data == each_name:
                    each_oc_data['Z_Port_Name'] = each_name
                    each_oc_data['Z_Port_ID'] = each_id

def Query_Optical_Route_Sheet_ID():
    Task_Name_ID_List = List_7013[1][5].split('-')
    Work_Sheet_Count = math.ceil(len(List_OC_Data) / 40)
    for work_sheet_num in range(Work_Sheet_Count):
        Work_Sheet_Name = '关于' + List_7013[1][0] + List_7013[1][1] + Task_Name_ID_List[1] + '二级光路申请' + '0' + str(work_sheet_num)
        Work_Sheet_Times = []
        Work_Sheet_Times.append(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
        Work_Sheet_Times.append((datetime.datetime.now() + datetime.timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S'))
        Work_Sheet_Times.append((datetime.datetime.now() + datetime.timedelta(days=4)).strftime('%Y-%m-%d %H:%M:%S'))
        Work_Sheet_Times.append((datetime.datetime.now() + datetime.timedelta(days=6)).strftime('%Y-%m-%d %H:%M:%S'))
        URL_Query_Work_Sheet_Exist = 'http://10.209.199.72:7112/irms/tasklistAction!waitedTaskAJAX.ilf'
        Form_Info_Encoded = 'processinstname=' + urllib.parse.quote_plus(Work_Sheet_Name)
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Query_Work_Sheet_Exist, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
        Response_Body = Response_Body.text
        Response_Body = json.loads(Response_Body)
        Exist_Work_Sheet = Response_Body['totalCount']
        if Exist_Work_Sheet == 0: # Create
            URL_Create_New_Work_Sheet_Step_1 = 'http://10.209.199.72:7112/irms/opticalSchedulingAction!init.ilf'
            Response_Body = requests.get(URL_Create_New_Work_Sheet_Step_1, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})
            Response_Body = bytes(Response_Body.text, encoding="utf-8")
            Response_Body = lxml.etree.HTML(Response_Body)
            List_Work_Sheet_Info = Response_Body.xpath('//input/@value')
            URL_Create_New_Work_Sheet_Step_2 = 'http://10.209.199.72:7112/irms/opticalSchedulingAction!submit.ilf?needSh=false'
            Form_Info_Encoded = 'ownerId=' + List_Work_Sheet_Info[1] + '&deptId=' + List_Work_Sheet_Info[2] + '&flowId=' + List_Work_Sheet_Info[3] + '&companyName=' + urllib.parse.quote_plus(List_Work_Sheet_Info[7]) + '&companyId=' + List_Work_Sheet_Info[8] + '&workitemId=' + List_Work_Sheet_Info[10] + '&activeName=' + List_Work_Sheet_Info[14] + '&formNo=' + List_Work_Sheet_Info[15] + '&title=' + urllib.parse.quote_plus(Work_Sheet_Name) + '&startFlag=' + List_Work_Sheet_Info[19] + '&ownerName=' + urllib.parse.quote_plus(List_Work_Sheet_Info[21]) + '&cellPhone=' +List_Work_Sheet_Info[22] + '&deptName=' + urllib.parse.quote_plus(List_Work_Sheet_Info[23]) + '&startTime=' + urllib.parse.quote_plus(Work_Sheet_Times[0]) + '&acceptTime=' + urllib.parse.quote_plus(Work_Sheet_Times[1]) + '&replyTime=' + urllib.parse.quote_plus(Work_Sheet_Times[2]) + '&requestTime=' + urllib.parse.quote_plus(Work_Sheet_Times[3]) + '&urgentDegree=' + urllib.parse.quote_plus('一般') + '&schedulingReason='
            Request_Lenth = str(len(Form_Info_Encoded))
            Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
            Response_Body = requests.post(URL_Create_New_Work_Sheet_Step_2, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})
            Repete_Start = (int(work_sheet_num) * 40)
            Repete_End = (int(work_sheet_num + 1) * 40)
            for oc_num in range(Repete_Start,Repete_End):
                List_OC_Data[oc_num]['Pro_ID'] = List_Work_Sheet_Info[3]
                if (oc_num + 1) == len(List_OC_Data):
                    break
        else:
            Repete_Start = (int(work_sheet_num) * 40)
            Repete_End = (int(work_sheet_num + 1) * 40)
            for oc_num in range(Repete_Start,Repete_End):
                List_OC_Data[oc_num]['Pro_ID'] = Response_Body['root'][0]['FLOW_ID']
                if (oc_num + 1) == len(List_OC_Data):
                    break

def Query_Integrate_Sheet_ID():

    Task_Name_ID_List = List_7013[1][5].split('-')

    Integrate_Sheet_Time = []
    Integrate_Sheet_Time.append(datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    Integrate_Sheet_Time.append((datetime.datetime.now() + datetime.timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S'))
    Integrate_Sheet_Time.append((datetime.datetime.now() + datetime.timedelta(days=4)).strftime('%Y-%m-%d %H:%M:%S'))

    Integrate_Sheet_Name = '关于' + List_7013[1][0] + List_7013[1][1] + Task_Name_ID_List[1] + '新建分纤箱入网申请'
    URL_Query_Integrate_Sheet_Exist = 'http://10.209.199.72:7112/irms/tasklistAction!waitedTaskAJAX.ilf'
    Form_Info_Encoded = 'processinstname=' + urllib.parse.quote_plus(Integrate_Sheet_Name)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Query_Integrate_Sheet_Exist, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})
    Response_Body = Response_Body.text
    Response_Body = json.loads(Response_Body)
    if Response_Body['totalCount'] == 0:

        print("新建管线入网工单", Integrate_Sheet_Name)
        URL_Create_Integrate_Sheet_Step_1 = 'http://10.209.199.72:7112/irms/pipelineresInAction!getProtectNameForCombox.ilf'
        Form_Info_Encoded = 'start=0&limit=10&paramName=' +  List_7013[1][4]
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Create_Integrate_Sheet_Step_1, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})
        Response_Body = json.loads(Response_Body.text)
        Project_Code_ID = Response_Body['root'][0]['INT_ID']

        URL_Create_Integrate_Sheet_Step_2 = 'http://10.209.199.72:7112/irms/pipelineresInAction!init.ilf'
        Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded'}
        Response_Body = requests.get(URL_Create_Integrate_Sheet_Step_2, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})
        Response_Body = bytes(Response_Body.text, encoding="utf-8")
        Response_Body = lxml.etree.HTML(Response_Body)
        List_Integrate_Sheet_Key = Response_Body.xpath('//input/@name')
        List_Integrate_Sheet_Value = Response_Body.xpath('//input/@value')
        Dic_Integrate_Sheet_Info = dict(zip(List_Integrate_Sheet_Key, List_Integrate_Sheet_Value))

        URL_Create_Integrate_Sheet_Step_3 = 'http://10.209.199.72:7112/irms/pipelineresInAction!saveDraft.ilf'
        Form_Info_Encoded = 'state=' + '&ownerId=' + str(Dic_Integrate_Sheet_Info['ownerId']) + '&deptId=' + str(Dic_Integrate_Sheet_Info['deptId']) + '&flowId=' + str(Dic_Integrate_Sheet_Info['flowId']) + '&id=' + '&pid=' + '&toppid=' + '&companyName=' + urllib.parse.quote_plus(Dic_Integrate_Sheet_Info['companyName']) + '&companyId=' + str(Dic_Integrate_Sheet_Info['companyId']) + '&currentstate=' + '&workitemId=' + '&formInfoId=' + '&prtFlowId=' + '&relationId=' + '&activeName=' + '&formNo=' + str(Dic_Integrate_Sheet_Info['formNo']) + '&title=' + urllib.parse.quote_plus(Integrate_Sheet_Name) + '&ownerName=' + urllib.parse.quote_plus(Dic_Integrate_Sheet_Info['ownerName']) + '&cellPhone=' + str(Dic_Integrate_Sheet_Info['cellPhone']) + '&deptName=' + urllib.parse.quote_plus(Dic_Integrate_Sheet_Info['deptName']) + '&startTime=' + urllib.parse.quote_plus(Integrate_Sheet_Time[0]) + '&acceptTime=' + urllib.parse.quote_plus(Integrate_Sheet_Time[1]) + '&replyTime=' + urllib.parse.quote_plus(Integrate_Sheet_Time[0]) + '&protectNo=' + str(Project_Code_ID) + '&protectName=' + List_7013[1][4] + '&taskNo=' + str(Task_Name_ID_List[0]) + '&taskName=' + urllib.parse.quote_plus(List_7013[1][5]) + '&remark='
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Create_Integrate_Sheet_Step_3, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})

        URL_Create_Integrate_Sheet_Step_4 = 'http://10.209.199.72:7112/irms/pipelineresInAction!init.ilf' +'?currentstate=0&flowId=' + Dic_Integrate_Sheet_Info['flowId']
        Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded'}
        Response_Body = requests.get(URL_Create_Integrate_Sheet_Step_4, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})
        Response_Body = bytes(Response_Body.text, encoding="utf-8")
        Response_Body = lxml.etree.HTML(Response_Body)
        List_Integrate_Sheet_Key = Response_Body.xpath('//input/@name')
        List_Integrate_Sheet_Value = Response_Body.xpath('//input/@value')
        Dic_Integrate_Sheet_Info = dict(zip(List_Integrate_Sheet_Key, List_Integrate_Sheet_Value))

        URL_Create_Integrate_Sheet_Step_5 = 'http://10.209.199.72:7112/irms/pipelineresInAction!submit.ilf'
        Form_Info_Encoded = 'state=' + '&ownerId=' + str(Dic_Integrate_Sheet_Info['ownerId']) + '&deptId=' + str(Dic_Integrate_Sheet_Info['deptId']) + '&flowId=' + str(Dic_Integrate_Sheet_Info['flowId']) + '&id=' + str(Dic_Integrate_Sheet_Info['id']) + '&pid=' + '&toppid=' + '&companyName=' + urllib.parse.quote_plus(Dic_Integrate_Sheet_Info['companyName']) + '&companyId=' + str(Dic_Integrate_Sheet_Info['companyId']) + '&currentstate=' + '&workitemId=' + '&formInfoId=' + '&prtFlowId=' + '&relationId=' + '&activeName=' + '&formNo=' + str(Dic_Integrate_Sheet_Info['formNo']) + '&title=' + urllib.parse.quote_plus(Integrate_Sheet_Name) + '&ownerName=' + urllib.parse.quote_plus(Dic_Integrate_Sheet_Info['ownerName']) + '&cellPhone=' + str(Dic_Integrate_Sheet_Info['cellPhone']) + '&deptName=' + urllib.parse.quote_plus(Dic_Integrate_Sheet_Info['deptName']) + '&startTime=' + urllib.parse.quote_plus(Integrate_Sheet_Time[0]) + '&acceptTime=' + urllib.parse.quote_plus(Integrate_Sheet_Time[1]) + '&replyTime=' + urllib.parse.quote_plus(Integrate_Sheet_Time[0]) + '&protectNo=' + str(Project_Code_ID) + '&protectName=' + List_7013[1][4] + '&taskNo=' + str(Task_Name_ID_List[0]) + '&taskName=' + urllib.parse.quote_plus(List_7013[1][5]) + '&remark='
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Create_Integrate_Sheet_Step_5, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Create, 'route': route_v_Create})

        Query_Integrate_Sheet_ID()

    elif Response_Body['totalCount'] != 0:
        print(Integrate_Sheet_Name, "工单Pro_ID", Response_Body['root'][0]['PROCESSINSTID'])
        for each_cs_data in List_CS_Data:
            each_cs_data['Pro_ID'] = Response_Body['root'][0]['PROCESSINSTID']

def Query_OC_Int_ID():
    List_Pro_ID = []
    for oc_data in List_OC_Data:
        List_Pro_ID.append(oc_data['Pro_ID'])
    Set_Pro_ID = set(List_Pro_ID)
    List_Pro_ID = list(Set_Pro_ID)

    for pro_id in List_Pro_ID:
        URL_Query_OC_Int_ID = 'http://10.209.199.72:7112/irms/opticOpenApplyAction!queryOptictemp.ilf?flowId='
        URL_Query_OC_Int_ID = URL_Query_OC_Int_ID + str(pro_id)
        Form_Info_Encoded = 'start=' + '0' + '&limit=' + '45'
        Request_Header = {"Host":"10.209.199.72:7112","Content-Type":"application/x-www-form-urlencoded"}
        Response_Body = requests.post(URL_Query_OC_Int_ID, headers=Request_Header, data=Form_Info_Encoded,  cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
        Response_Body = Response_Body.text
        Response_Body = json.loads(Response_Body)
        for oc_info_in_response in Response_Body['root']:
            for oc_data in List_OC_Data:
                if oc_data['OC_Name'] == oc_info_in_response['opticname']:
                    oc_data['Int_ID'] = oc_info_in_response['intId']
                    break


def Execute_Push_Box():
    URL_Push_Box = 'http://10.209.199.74:8120/igisserver_osl/rest/ResourceController/resourcesUpdate?isUpdate=move'
    Form_Info_Head = '<xmldata mode="SinglePointEditMode"><mc type="ziyuandian">'
    Form_Info_Tail = '</mc></xmldata>'
    List_Form_Info_Body = []
    for each_box_data in List_Box_Data:
        Form_Info_Body = '<mo group="1"><fv k="LATITUDE" v="' + str(each_box_data['Latitude']) + '"/><fv k="INT_ID" v="' + str(each_box_data['ResPoint_ID']) + '"/><fv k="ZH_LABEL" v="' + each_box_data['ResPoint_Name'] + '"/><fv k="LONGITUDE" v="' + str(each_box_data['Longitude']) + '"/></mo>'
        List_Form_Info_Body.append(Form_Info_Body)
    Form_Info_Body = ''.join(List_Form_Info_Body)
    Form_Info = Form_Info_Head + Form_Info_Body + Form_Info_Tail
    Form_Info_Encoded = 'xml=' + urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Push_Box, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    Push_Result = Response_Body.xpath('//@loaded')
    print('P1-推箱子-{}'.format(Push_Result[0]))

def Execute_Generate_Support_Segment():
    URL_Generate_Support_Segment = 'http://10.209.199.74:8120/igisserver_osl/rest/ResourceController/resourcesAdd'
    Form_Info_Head = '<xmldata mode="PipeLineAddMode"><mc type="yinshangduan">'
    Form_Info_Tail = '</mc></xmldata>'
    List_Form_Info_Body = []
    for ocs_num in List_CS_Data:
        Form_Info_Body = '<mo group="1" ax="'+str(ocs_num['A_Longitude'])+'" ay="'+str(ocs_num['A_Latitude'])+'" zx="'+str(ocs_num['Z_Longitude'])+'" zy="'+str(ocs_num['Z_Latitude'])+'"><fv k="CITY_ID" v="'+str(ocs_num['City_ID'])+'"/><fv k="COUNTY_ID" v="'+str(ocs_num['County_ID'])+'"/><fv k="RELATED_SYSTEM" v="'+str(ocs_num['Support_Sys_ID'])+'"/><fv k="A_OBJECT_ID" v="'+str(ocs_num['A_ResPoint_ID'])+'"/><fv k="ZH_LABEL" v="'+str(ocs_num['A_Box_Name'])+'资源点-'+str(ocs_num['Z_Box_Name'])+'资源点引上段'+'"/><fv k="MAINTAINOR" v="'+str(ocs_num['DQS_Maintainer_ID'])+'"/><fv k="M_LENGTH" v="'+str(ocs_num['Length'])+'"/><fv k="STATUS" v="'+str(ocs_num['Life_Cycle'])+'"/><fv k="QUALITOR_COUNTY" v="'+str(ocs_num['DQS_County_ID'])+'"/><fv k="QUALITOR" v="'+str(ocs_num['DQS_ID'])+'"/><fv k="QUALITOR_PROJECT" v="'+str(ocs_num['DQS_Project_ID'])+'"/><fv k="OWNERSHIP" v="'+str(ocs_num['Owner_Type'])+'"/><fv k="RES_OWNER" v="'+str(ocs_num['Owner_Name'])+'"/><fv k="A_OBJECT_TYPE" v="'+str(ocs_num['A_ResPoint_Type_ID'])+'"/><fv k="INT_ID" v="new-27991311"/><fv k="TASK_NAME" v="'+str(ocs_num['Task_Name_ID'])+'"/><fv k="PROJECT_CODE" v="'+str(ocs_num['Project_Code_ID'])+'"/><fv k="RESOURCE_LOCATION" v="'+str(ocs_num['Field_Type'])+'"/><fv k="Z_OBJECT_TYPE" v="'+str(ocs_num['Z_ResPoint_Type_ID'])+'"/><fv k="Z_OBJECT_ID" v="'+str(ocs_num['Z_ResPoint_ID'])+'"/><fv k="SYSTEM_LEVEL" v="'+str(ocs_num['Business_Level'])+'<fv k="PRO_TASK_ID" v="'+str(ocs_num['Pro_ID'])+'"/></mo>'
        List_Form_Info_Body.append(Form_Info_Body)
    Form_Info_Body = ''.join(List_Form_Info_Body)
    Form_Info = Form_Info_Head + Form_Info_Body + Form_Info_Tail
    Form_Info_Encoded = 'empname=' + urllib.parse.quote_plus('白云鹏') + '&xml=' + urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Generate_Support_Segment, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Generate_Support_Segment_State = Response_Body.xpath('//@loaded')
    print('P2-引上段建立-{}'.format(List_Generate_Support_Segment_State[0]))

def Execute_Generate_Cable_Segment():
    URL_Generate_Cable_Segment = 'http://10.209.199.74:8120/igisserver_osl/rest/ResourceController/resourcesAdd?coreNamingRules=0'
    Form_Info_Head = '<xmldata mode="FibersegAddMode"><mc type="guanglanduan">'
    Form_Info_Tail = '</mc></xmldata>'
    List_Form_Info_Body = []
    for ocs_num in List_CS_Data:
        Form_Info_Body = '<mo group="1" ax="'+str(ocs_num['A_Longitude'])+'" ay="'+str(ocs_num['A_Latitude'])+'" zx="'+str(ocs_num['Z_Longitude'])+'" zy="'+str(ocs_num['Z_Latitude'])+'"><fv k="QUALITOR_PROJECT" v="'+str(ocs_num['DQS_Project_ID'])+'"/><fv k="RES_OWNER" v="'+str(ocs_num['Owner_Name'])+'"/><fv k="RELATED_SYSTEM" v="'+str(ocs_num['Cable_Sys_ID'])+'"/><fv k="QUALITOR" v="'+str(ocs_num['DQS_ID'])+'"/><fv k="ZH_LABEL" v="'+str(ocs_num['A_Box_Name'])+'资源点-'+str(ocs_num['Z_Box_Name'])+'资源点'+'"/><fv k="STATUS" v="'+str(ocs_num['Life_Cycle'])+'"/><fv k="FIBER_TYPE" v="2"/><fv k="INT_ID" v="new-27991311"/><fv k="A_OBJECT_TYPE" v="'+str(ocs_num['A_ResPoint_Type_ID'])+'"/><fv k="Z_OBJECT_ID" v="'+str(ocs_num['Z_ResPoint_ID'])+'"/><fv k="A_OBJECT_ID" v="'+str(ocs_num['A_ResPoint_ID'])+'"/><fv k="Z_OBJECT_TYPE" v="'+str(ocs_num['Z_ResPoint_Type_ID'])+'"/><fv k="M_LENGTH" v="'+str(ocs_num['Length'])+'"/><fv k="CITY_ID" v="'+str(ocs_num['City_ID'])+'"/><fv k="FIBER_NUM" v="'+str(ocs_num['Width'])+'"/><fv k="MAINTAINOR" v="'+str(ocs_num['DQS_Maintainer_ID'])+'"/><fv k="WIRE_SEG_TYPE" v="GYTA-'+str(ocs_num['Width'])+'"/><fv k="SERVICE_LEVEL" v="14"/><fv k="COUNTY_ID" v="'+str(ocs_num['County_ID'])+'"/><fv k="PROJECTCODE" v="'+str(ocs_num['Project_Code_ID'])+'"/><fv k="TASK_NAME" v="'+str(ocs_num['Task_Name_ID'])+'"/><fv k="OWNERSHIP" v="'+str(ocs_num['Owner_Type'])+'"/><fv k="QUALITOR_COUNTY" v="'+str(ocs_num['DQS_County_ID'])+'<fv k="PRO_TASK_ID" v="'+str(ocs_num['Pro_ID'])+'"/></mo>'
        List_Form_Info_Body.append(Form_Info_Body)
    Form_Info_Body = ''.join(List_Form_Info_Body)
    Form_Info = Form_Info_Head + Form_Info_Body + Form_Info_Tail
    Form_Info_Encoded = 'empname=' + urllib.parse.quote_plus('白云鹏') + '&xml=' + urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Generate_Cable_Segment, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Generate_Cable_Segment_State = Response_Body.xpath('//@loaded')
    print('P3-光缆段建立-{}'.format(List_Generate_Cable_Segment_State[0]))

def Execute_Cable_Lay(Para_List_CS_Data):
    URL_Cable_Lay = 'http://10.209.199.74:8120/igisserver_osl/rest/optCabLayInspur/saveFiberSegM1'
    Form_Info = '<xmldata><fiberseg id="'+str(Para_List_CS_Data['Cable_Seg_ID'])+'" aid="'+str(Para_List_CS_Data['A_ResPoint_ID'])+'" zid="'+str(Para_List_CS_Data['Z_ResPoint_ID'])+'"/><cablays><cablay id="'+str(Para_List_CS_Data['Support_Seg_ID'])+'" type="yinshangduan" name="'+str(Para_List_CS_Data['Support_Seg_Name'])+'" aid="'+str(Para_List_CS_Data['A_ResPoint_ID'])+'" zid="'+str(Para_List_CS_Data['Z_ResPoint_ID'])+'"/></cablays></xmldata>'
    Form_Info_Encoded = 'xml='+urllib.parse.quote_plus(Form_Info)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Cable_Lay, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Cable_Lay_State = Response_Body.xpath('//@msg')
    print('P4-敷设-{}-{}'.format(List_Cable_Lay_State[0], Para_List_CS_Data['Cable_Seg_Name']))

def Execute_Generate_ODM(Para_List_Box_Data):
    URL_Generate_ODM = 'http://10.209.199.74:8120/nxapp/room/editODMData.ilf'
    Form_Info = '<params><odm id="" rowflag="+" rownum="1" colflag="+" colnum="12"><attribute module_rowno="1" rowno="'+str(Para_List_Box_Data['ODM_Rows'])+'" columnno="12" terminal_arr="0" maintain_county="'+str(Para_List_Box_Data['County_ID'])+'" maintain_city="'+str(Para_List_Box_Data['City_ID'])+'" structure_id="'+str(Para_List_Box_Data['ResPoint_ID'])+'" structure_type="'+str(Para_List_Box_Data['ResPoint_Type_ID'])+'" related_rack="'+str(Para_List_Box_Data['Box_ID'])+'" related_type="'+str(Para_List_Box_Data['Box_Type_ID'])+'" status="8" model="odm"/></odm></params>'
    Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
    Form_Info_Encoded = "params=" + urllib.parse.quote_plus(Form_Info) + "&model=odm&" +  "lifeparams=" + urllib.parse.quote_plus(Form_Info_Tail)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Generate_ODM, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = urllib.parse.unquote(bytes(Response_Body.text, encoding="utf-8"))
    Response_Body = lxml.etree.HTML(Response_Body)
    List_ODM_ID = Response_Body.xpath('//@int_id')
    Para_List_Box_Data['ODM_ID'] = List_ODM_ID[0]
    print('P5-ODM-{}-in-{}'.format(List_ODM_ID[0], Para_List_Box_Data['Box_Name']))

def Execute_Generate_Tray(Para_List_Box_Data):
    URL_Generate_Tray = "http://10.209.199.74:8120/nxapp/room/editTray_sx.ilf"
    Form_Info = '<params model="tray"><obj related_rack="'+str(Para_List_Box_Data['Box_ID'])+'" related_type="'+str(Para_List_Box_Data['Box_Type_ID'])+'" structure_id="'+str(Para_List_Box_Data['ResPoint_ID'])+'" structure_type="'+str(Para_List_Box_Data['ResPoint_Type_ID'])+'" deviceshelf_id="'+str(Para_List_Box_Data['ODM_ID'])+'" tray_no="1" tray_num="'+str(Para_List_Box_Data['Tray_Count'])+'" row_count="1" col_count="12" int_id=""/></params>'
    Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
    Form_Info_Encoded = "params=" + urllib.parse.quote_plus(Form_Info) + "&model=odm&" +  "lifeparams=" + urllib.parse.quote_plus(Form_Info_Tail)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Generate_Tray, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = urllib.parse.unquote(bytes(Response_Body.text, encoding="utf-8"))
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Terminal_IDs = Response_Body.xpath('//terminal//@int_id')
    print('P6-Terminal_IDs_Count-{}-{}'.format(len(List_Terminal_IDs), Para_List_Box_Data['Box_Name']))
    Para_List_Box_Data['Terminal_IDs'] = List_Terminal_IDs

def Execute_Termination(Para_List_Box_Data):

    if Para_List_Box_Data['1FS_Count'] == 0:
        termination_datas = []
        for termination_num in range(Para_List_Box_Data['Termination_Count']):
            termination_data = '<param fiber_id="'+str(Para_List_Box_Data['Termination_Fiber_IDs'][termination_num])+'" z_equ_id="'+str(Para_List_Box_Data['Box_ID'])+'" z_equ_type="'+str(Para_List_Box_Data['Box_Type_Short'])+'" z_port_id="'+str(Para_List_Box_Data['Terminal_IDs'][termination_num])+'" a_equ_id="" a_port_id="" room_id="'+str(Para_List_Box_Data['ResPoint_ID'])+'"/>'
            termination_datas.append(termination_data)
        Form_Info_Body = ''.join(termination_datas)

    elif Para_List_Box_Data['1FS_Count'] != 0:
        termination_datas = []
        for termination_num in range(int(len(Para_List_Box_Data['Direct_Melt_Count']))):
            termination_data = '<param fiber_id="'+str(Para_List_Box_Data['Direct_Melt_Count'][termination_num][3])+'" a_equ_id="'+str(Para_List_Box_Data['Box_ID'])+'" a_equ_type="'+Para_List_Box_Data['Box_Type_Short']+'" a_port_id="'+str(Para_List_Box_Data['Direct_Melt_Count'][termination_num][1])+'" z_equ_id="" z_port_id="" room_id="'+Para_List_Box_Data['ResPoint_ID']+'"/>'
            termination_datas.append(termination_data)
        Form_Info_Body = ''.join(termination_datas)

    URL_Termination = 'http://10.209.199.74:8120/igisserver_osl/rest/jumperandfiber/saveFiber'
    Form_Info_Head = '<params>'
    Form_Info_Butt = '</params>'
    Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
    Form_Info_Encoded = 'params=' + urllib.parse.quote_plus(Form_Info_Head + Form_Info_Body + Form_Info_Butt) + '&lifeparams=' + urllib.parse.quote_plus(Form_Info_Tail)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Termination, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Termination_State = Response_Body.xpath('//@msg')
    print('P7-{}-{}'.format(Para_List_Box_Data['Box_Name'], List_Termination_State[0]))

def Execute_Direct_Melt(Para_List_Box_Data):
    if Para_List_Box_Data['1FS_Count'] == 0:

        for cable_seg_data in List_CS_Data:
            if Para_List_Box_Data['Box_Name'] == cable_seg_data['Z_Box_Name']:
                List_UL_CS_Fiber_IDs = cable_seg_data['CS_Fiber_IDs']
            if Para_List_Box_Data['Box_Name'] == cable_seg_data['A_Box_Name']:
                List_DL_CS_Fiber_IDs = cable_seg_data['CS_Fiber_IDs']

        URL_Direct_Melt = 'http://10.209.199.74:8120/igisserver_osl/rest/fibercorekiss/fiberKiss'

        Form_Info_Head = '<params respoint_id="'+Para_List_Box_Data['ResPoint_ID']+'" respoint_type="'+str(Para_List_Box_Data['ResPoint_Type_ID'])+'" parent_rack_id="null" parent_rack_type="null">'
        each_direct_melt_datas = []
        for each_direct_melt_data_num in range(Para_List_Box_Data['Direct_Melt_Count']):
            each_direct_melt_data = '<param a_fiber_id="'+str(List_UL_CS_Fiber_IDs[each_direct_melt_data_num + Para_List_Box_Data['Direct_Melt_Start'] - 1])+'" z_fiber_id="'+str(List_DL_CS_Fiber_IDs[each_direct_melt_data_num])+'"/>'
            each_direct_melt_datas.append(each_direct_melt_data)
        Form_Info_Body = ''.join(each_direct_melt_datas)
        Form_Info_Butt = '</params>'
        Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
        Form_Info_Encoded = 'params=' + urllib.parse.quote_plus(Form_Info_Head + Form_Info_Body + Form_Info_Butt) + '&lifeparams=' + urllib.parse.quote_plus(Form_Info_Tail)
        Request_Lenth = str(len(Form_Info_Encoded))
        Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
        Response_Body = requests.post(URL_Direct_Melt, data=Form_Info_Encoded, headers=Request_Header)
        Response_Body = bytes(Response_Body.text, encoding="utf-8")
        Response_Body = lxml.etree.HTML(Response_Body)
        List_Direct_Melt_State = Response_Body.xpath('//@msg')
        print('P8-{}-{}'.format(Para_List_Box_Data['Box_Name'], List_Direct_Melt_State[0]))

    elif Para_List_Box_Data['1FS_Count'] != 0:
        print('P8-{}-是一级分纤箱,不涉及直熔'.format(Para_List_Box_Data['Box_Name']))

def Execute_Generate_Optical_Circut(Para_List_OC_Data):
    URL_Generate_Optical_Circut = 'http://10.209.199.72:7112/irms/opticOpenAction!saveOptictemp.ilf'
    Form_Info_Encoded = 'proId=' + str(Para_List_OC_Data['Pro_ID']) + '&cityId=' + str(Para_List_OC_Data['City_ID']) + '&countyId=' + str(Para_List_OC_Data['County_ID']) + '&aportequname=' + urllib.parse.quote_plus(Para_List_OC_Data['A_Box_Name']) + '&aportequtype=' + str(Para_List_OC_Data['A_Box_Type_ID']) + '&aportequid=' + str(Para_List_OC_Data['A_Box_ID']) + '&asitename=' + urllib.parse.quote_plus(Para_List_OC_Data['A_ResPoint_Name']) + '&asitetype=' + str(Para_List_OC_Data['A_ResPoint_Type_ID']) + '&asiteid=' + str(Para_List_OC_Data['A_ResPoint_ID']) + '&aequname=' + urllib.parse.quote_plus(Para_List_OC_Data['A_POS_Name']) + '&aequid=' + str(Para_List_OC_Data['A_POS_ID']) + '&aequtype=' + str(Para_List_OC_Data['AEquType']) + '&ajoinName=' + urllib.parse.quote_plus(Para_List_OC_Data['AJoinName']) + '&zportequname=' + urllib.parse.quote_plus(Para_List_OC_Data['Z_Box_Name']) + '&zportequtype=' + str(Para_List_OC_Data['Z_Box_Type_ID']) + '&zportequid=' + str(Para_List_OC_Data['Z_Box_ID']) + '&zsitename=' + urllib.parse.quote_plus(Para_List_OC_Data['Z_ResPoint_Name']) + '&zsitetype=' + str(Para_List_OC_Data['Z_ResPoint_Type_ID']) + '&zsiteid=' + str(Para_List_OC_Data['Z_ResPoint_ID']) + '&zequname=' + urllib.parse.quote_plus(Para_List_OC_Data['Z_POS_Name']) + '&zequid=' + str(Para_List_OC_Data['Z_POS_ID']) + '&zequtype=' + str(Para_List_OC_Data['ZEquType']) + '&zjoinName=' + urllib.parse.quote_plus(Para_List_OC_Data['ZJoinName']) + '&opticname=' + urllib.parse.quote_plus(Para_List_OC_Data['OC_Name']) + '&chengzaiyewu=' + urllib.parse.quote_plus(Para_List_OC_Data['Business_Name']) + '&sxbusstype=' + urllib.parse.quote_plus(Para_List_OC_Data['SXBussType']) + '&bakinfo=' + urllib.parse.quote_plus('无') + '&bussType=' + str(Para_List_OC_Data['BussType']) + '&apptype=' + str(Para_List_OC_Data['AppType']) + '&serviceLevel=' + str(Para_List_OC_Data['ServiceLevel']) + '&fiberCount=' + str(Para_List_OC_Data['FiberCount']) + '&qualityGc=' + urllib.parse.quote_plus(Para_List_OC_Data['DQS_Project']) + '&qualityGcId=' + str(Para_List_OC_Data['DQS_Project_ID']) + '&qualityDs=' + urllib.parse.quote_plus(Para_List_OC_Data['DQS']) + '&qualityDsId=' + str(Para_List_OC_Data['DQS_ID']) + '&qualityQx=' + urllib.parse.quote_plus(Para_List_OC_Data['DQS_County']) + '&qualityQxId=' + str(Para_List_OC_Data['DQS_County_ID']) + '&maintain=' + urllib.parse.quote_plus(Para_List_OC_Data['DQS_Maintainer']) + '&maintainId=' + str(Para_List_OC_Data['DQS_Maintainer_ID']) + '&projectCodeName=' + urllib.parse.quote_plus(Para_List_OC_Data['Project_Code']) + '&projectCodeId=' + str(Para_List_OC_Data['Project_Code_ID']) + '&taskName=' + urllib.parse.quote_plus(Para_List_OC_Data['Task_Name']) + '&taskNameid=' + str(Para_List_OC_Data['Task_Name_ID'])
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host':'10.209.199.72:7112', 'Content-Type': 'application/x-www-form-urlencoded', 'Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Generate_Optical_Circut, data=Form_Info_Encoded, headers=Request_Header, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
    Response_Body = Response_Body.text
    Response_Body = Response_Body.replace('success:true','"success":true')
    Response_Body = Response_Body.replace('mesg','"mesg"')
    Response_Body = Response_Body.replace('lyxx','"lyxx"')
    Response_Body = Response_Body.replace('lytypexx','"lytypexx"')
    Response_Body = Response_Body.replace('workorder','"workorder"')
    Response_Body = Response_Body.replace('\'','\"')
    Response_Body = Response_Body.replace('x\"t','xt')
    Response_Body = Response_Body.replace('r\"i','ri')
    Response_Body = Response_Body.replace('\"\"','\"')
    Response_Body = json.loads(Response_Body)
    print('P9-{}-{}'.format(Response_Body['mesg'] ,Para_List_OC_Data['OC_Name']))

def Execute_Transmission_Design(Para_List_OC_Data):
    if Para_List_OC_Data['A_Box_Name'] != Para_List_OC_Data['Z_Box_Name']:# 业务光路

        # 通路占用开始
        URL_Query_Path_IDs = 'http://10.209.199.72:7112/irms/opticOpenDesignAction!Searchpath4page.ilf?'
        Query_Detail = 'flowId=' + str(Para_List_OC_Data['Pro_ID']) + '&workorid=' + str(Para_List_OC_Data['Int_ID']) + '&aobject=' + str(Para_List_OC_Data['A_Box_ID']) + '&zobject=' + str(Para_List_OC_Data['Z_Box_ID']) + '&aobjectType=' + str(Para_List_OC_Data['A_Box_Type_ID']) + '&zobjectType=' + str(Para_List_OC_Data['Z_Box_Type_ID']) + '&limit=20&start=1'
        URL_Query_Path_IDs = URL_Query_Path_IDs + Query_Detail
        Response_Body = requests.get(URL_Query_Path_IDs, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
        Response_Body = lxml.etree.HTML(Response_Body.text)

        # 处理重复通路的系统bug开始
        List_Path_IDs = Response_Body.xpath('//line/@id')
        List_FiberNo = Response_Body.xpath('//line/@fiberno')
        Dic_Paths = dict(zip(List_FiberNo,List_Path_IDs))
        List_Path_IDs = []
        for path in Dic_Paths.values():
            List_Path_IDs.append(path)
        Para_List_OC_Data['Occupy_Path'] = List_Path_IDs[(List_OC_Path_Pointer[Para_List_OC_Data['Z_Box_Name']])]
        # 处理重复通路的系统bug结束
        URL_Occupy_Path = 'http://10.209.199.72:7112/irms/opticOpenDesignAction!occupyPath.ilf?'
        Query_Detail = 'flowId=' + str(Para_List_OC_Data['Pro_ID']) + '&workorid=' + str(Para_List_OC_Data['Int_ID']) + '&pathid=' + str(Para_List_OC_Data['A_Box_ID']) + ';' + str(Para_List_OC_Data['Z_Box_ID']) + '-' + '&dpath=' + str(Para_List_OC_Data['A_Box_ID']) + ';' + str(Para_List_OC_Data['Z_Box_ID']) + '&ids=' + str(Para_List_OC_Data['Occupy_Path'])
        URL_Occupy_Path = URL_Occupy_Path + Query_Detail
        Response_Body = requests.get(URL_Occupy_Path, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
        Response_Body = lxml.etree.HTML(Response_Body.text)
        Occupy_Result = Response_Body.xpath('//@success')

        print('P10通路占用-{}-{}'.format(Occupy_Result[0], Para_List_OC_Data['OC_Name']))
        # 通路占用结束

        # 端口配置开始
        URL_Port_Config = 'http://10.209.199.72:7112/irms/opticOpenDesignAction!updateworkernew.ilf'
        Form_Info_Encoded = 'flowId=' + str(Para_List_OC_Data['Pro_ID']) + '&workorid=' + str(Para_List_OC_Data['Int_ID']) + '&aobjecttype=' + str(Para_List_OC_Data['A_ResPoint_Type_ID']) + "&zobjecttype=" + str(Para_List_OC_Data['Z_ResPoint_Type_ID']) + '&aobjectid=' + str(Para_List_OC_Data['A_ResPoint_ID']) + '&zobjectid=' + str(Para_List_OC_Data['Z_ResPoint_ID']) + '&aobjectname=' + urllib.parse.quote_plus(str(Para_List_OC_Data['A_ResPoint_Name'])) + '&zobjectname=' + urllib.parse.quote_plus(str(Para_List_OC_Data['Z_ResPoint_Name'])) + '&aneid=' + str(Para_List_OC_Data['A_POS_ID']) + "&zneid=" + str(Para_List_OC_Data['Z_POS_ID']) + '&anename=' + urllib.parse.quote_plus(str(Para_List_OC_Data['A_POS_Name'])) + '&znename=' + urllib.parse.quote_plus(str(Para_List_OC_Data['Z_POS_Name'])) + '&aportid=' + str(Para_List_OC_Data['A_Port_ID']) + '&zportid=' + str(Para_List_OC_Data['Z_Port_ID']) + '&aportname=' + urllib.parse.quote_plus(str(Para_List_OC_Data['A_Port_Name'])) + '&zportname=' + urllib.parse.quote_plus(str(Para_List_OC_Data['Z_Port_Name'])) + '&isgenerateFiber=' + '1' + '&apptype=' + str(Para_List_OC_Data['AppType']) + '&anetype=' + str(Para_List_OC_Data['AEquType']) + "&znetype=" + str(Para_List_OC_Data['ZEquType'])
        Request_Header = {"Host":"10.209.199.72:7112","Content-Type":"application/x-www-form-urlencoded"}
        Response_Body = requests.post(URL_Port_Config,headers = Request_Header, data=Form_Info_Encoded, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
        Response_Body = Response_Body.text
        Response_Body = Response_Body.replace('success','"success"')
        Response_Body = Response_Body.replace('mesg','"mesg"')
        Response_Body = Response_Body.replace('detail','"detail"')
        Response_Body = Response_Body.replace('\'','\"')
        Response_Body = json.loads(Response_Body)
        print('P10配置端口-{}-{}'.format(Response_Body['mesg'] ,Para_List_OC_Data['OC_Name']))
        # 处理重复通路的系统bug开始
        List_OC_Path_Pointer[Para_List_OC_Data['Z_Box_Name']] += 1
        # 处理重复通路的系统bug结束
        # 端口配置结束

    elif Para_List_OC_Data['A_Box_Name'] == Para_List_OC_Data['Z_Box_Name']:# 尾纤光路

        # 跳纤开始
        URL_Save_Jumper = 'http://10.209.199.74:8120/igisserver_osl/rest/jumperandfiber/saveJumper'
        Form_Info_1 = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
        Form_Info_2 = '<params><param a_site_type="' + str(Para_List_OC_Data['A_ResPoint_Type_ID']) + '" a_site_id="' + str(Para_List_OC_Data['A_ResPoint_ID']) + '" a_equ_type="pos" a_equ_id="' + str(Para_List_OC_Data['A_POS_ID']) + '" a_port_type="pos_port" a_port_id="' + str(Para_List_OC_Data['A_Port_ID']) + '" z_site_type="'+ str(Para_List_OC_Data['Z_ResPoint_Type_ID']) + '" z_site_id="' + str(Para_List_OC_Data['Z_ResPoint_ID']) + '" z_equ_type="pos" z_equ_id="' + str(Para_List_OC_Data['Z_POS_ID']) + '" z_port_type="pos_port" z_port_id="' + str(Para_List_OC_Data['Z_Port_ID']) + '"/></params>'
        Form_Info_Encoded = 'params=' + urllib.parse.quote_plus(Form_Info_2) + '&lifeparams=' + urllib.parse.quote_plus(Form_Info_1)
        Request_Header = {"Host":"10.209.199.74:8120","Content-Type":"application/x-www-form-urlencoded"}
        Response_Body = requests.post(URL_Save_Jumper, headers=Request_Header, data=Form_Info_Encoded)
        Response_Body = bytes(Response_Body.text, encoding="utf-8")
        Response_Body = lxml.etree.HTML(Response_Body)
        List_Jumper_State = Response_Body.xpath('//@msg')
        print('P10跳纤-{}-{}'.format(List_Jumper_State[0], Para_List_OC_Data['OC_Name']))
        # 跳纤结束

        # 端口配置开始
        URL_Port_Config = 'http://10.209.199.72:7112/irms/opticOpenDesignAction!saveweixianpeizhi.ilf'
        Form_Info_Encoded = 'flowId=' + str(Para_List_OC_Data['Pro_ID']) + '&workorid=' + str(Para_List_OC_Data['Int_ID'])+'&aobjecttype=' + str(Para_List_OC_Data['A_ResPoint_Type_ID']) + '&zobjecttype=' + str(Para_List_OC_Data['Z_ResPoint_Type_ID']) + '&aobjectid=' + str(Para_List_OC_Data['A_ResPoint_ID']) + '&zobjectid=' + str(Para_List_OC_Data['Z_ResPoint_ID']) + '&aobjectname=' + urllib.parse.quote_plus(Para_List_OC_Data['A_ResPoint_Name']) + '&zobjectname=' + urllib.parse.quote_plus(Para_List_OC_Data['Z_ResPoint_Name']) + '&aneid=' + str(Para_List_OC_Data['A_POS_ID']) + '&zneid=' + str(Para_List_OC_Data['Z_POS_ID']) + '&anename=' + urllib.parse.quote_plus(Para_List_OC_Data['A_POS_Name']) + '&znename=' + urllib.parse.quote_plus(Para_List_OC_Data['Z_POS_Name']) + '&aportid=' + str(Para_List_OC_Data['A_Port_ID']) + '&zportid=' + str(Para_List_OC_Data['Z_Port_ID']) + '&aportname=' + urllib.parse.quote_plus(Para_List_OC_Data['A_Port_Name']) + '&zportname=' + urllib.parse.quote_plus(Para_List_OC_Data['Z_Port_Name']) + '&apptype=' + str(Para_List_OC_Data['AppType']) + '&anetype=' + str(Para_List_OC_Data['AEquType']) + '&znetype=' + str(Para_List_OC_Data['ZEquType']) + '&fibercount=' + '1'
        Request_Header = {"Host":"10.209.199.72:7112","Content-Type":"application/x-www-form-urlencoded"}
        Response_Body = requests.post(URL_Port_Config,headers = Request_Header, data=Form_Info_Encoded, cookies={'JSESSIONIRMS': Jsessionirms_v_Run, 'route': route_v_Run})
        Response_Body = Response_Body.text
        Response_Body = Response_Body.replace('success','"success"')
        Response_Body = Response_Body.replace('mesg','"mesg"')
        Response_Body = Response_Body.replace('detail','"detail"')
        Response_Body = Response_Body.replace('\'','\"')
        Response_Body = Response_Body.replace('""detail"','"detail')
        Response_Body = json.loads(Response_Body)
        print('P10配置端口-{}-{}'.format(Response_Body['mesg'] ,Para_List_OC_Data['OC_Name']))
        # 端口配置结束

def Execute_Termination_2nd(Para_List_Box_Data):

    if Para_List_Box_Data['1FS_Count'] == 0:
        pass

    elif Para_List_Box_Data['1FS_Count'] != 0:
        termination_datas = []
        for termination_num in range(int(len(Para_List_Box_Data['Direct_Melt_Start']))):
            termination_data = '<param fiber_id="'+str(Para_List_Box_Data['Direct_Melt_Start'][termination_num][3])+'" a_equ_id="'+str(Para_List_Box_Data['Box_ID'])+'" a_equ_type="'+Para_List_Box_Data['Box_Type_Short']+'" a_port_id="'+str(Para_List_Box_Data['Direct_Melt_Start'][termination_num][1])+'" z_equ_id="" z_port_id="" room_id="'+Para_List_Box_Data['ResPoint_ID']+'"/>'
            termination_datas.append(termination_data)
        Form_Info_Body = ''.join(termination_datas)

    URL_Termination = 'http://10.209.199.74:8120/igisserver_osl/rest/jumperandfiber/saveFiber'
    Form_Info_Head = '<params>'
    Form_Info_Butt = '</params>'
    Form_Info_Tail = '<params><param key="pro_task_id" value=""/><param key="status" value="8"/><param key="photo" value="null"/><param key="isvirtual" value="0"/><param key="virtualtype" value=""/></params>'
    Form_Info_Encoded = 'params=' + urllib.parse.quote_plus(Form_Info_Head + Form_Info_Body + Form_Info_Butt) + '&lifeparams=' + urllib.parse.quote_plus(Form_Info_Tail)
    Request_Lenth = str(len(Form_Info_Encoded))
    Request_Header = {'Host': '10.209.199.74:8120','Content-Type': 'application/x-www-form-urlencoded','Content-Length': Request_Lenth}
    Response_Body = requests.post(URL_Termination, data=Form_Info_Encoded, headers=Request_Header)
    Response_Body = bytes(Response_Body.text, encoding="utf-8")
    Response_Body = lxml.etree.HTML(Response_Body)
    List_Termination_State = Response_Body.xpath('//@msg')
    print('P11-{}-{}'.format(Para_List_Box_Data['Box_Name'], List_Termination_State[0]))

def Excute_Update_1(Para_List_CS_Data):
    Dic_Response = []
    URL_Query_OpticalCable = 'http://10.209.199.74:8120/igisserver_osl/rest/generalSaveOrGet/generalGet'
    XML_Info_Encoded = 'xml=' + urllib.parse.quote_plus('''<request><query mc="guanglanduan" ids="" where="1=1 AND ZH_LABEL LIKE '%''' + Para_List_CS_Data['A_Box_Name'] + '资源点-' + Para_List_CS_Data['Z_Box_Name'] + '资源点' + '''%'" returnfields="INT_ID,ZH_LABEL,ALIAS,CITY_ID,COUNTY_ID,STATUS,PRO_TASK_ID,CJ_TASK_ID,A_EQUIP_TYPE,CJ_STATUS,A_EQUIP_ID,A_ROOM_ID,A_OBJECT_TYPE,A_OBJECT_ID,Z_EQUIP_TYPE,Z_EQUIP_ID,Z_ROOM_ID,Z_OBJECT_TYPE,Z_OBJECT_ID,SERVICE_LEVEL,C_LENGTH,M_LENGTH,FIBER_NUM,RELATED_SYSTEM,WIRE_SEG_TYPE,DIRECTION,IS_ALTER,FIBER_TYPE,DIA,GT_VERSION,VENDOR,OPTI_CABLE_TYPE,MAINT_DEP,MAINT_MODE,LAY_TYPE,RELATED_IS_AREA,IS_WRONG,MAINT_STATE,AREA_LEVEL,WRONG_INFO,SYS_VERSION,OWNERSHIP,PHONE_NO,TIME_STAMP,RES_OWNER,PRODUCE_DATE,PROJECTCODE,BUSINESS,TASK_NAME,ASSENT_NO,SERVICER,RUWANG_DATE,STUFF,BUILDER,TUIWANG_DATE,QUALITOR_PROJECT,PROJECT_ID,QUALITOR,QUALITOR_COUNTY,MAINTAINOR,QRCODE,SEG_NO,REMARK,INDEX_IN_BRANCH,IS_COMPLETE_LAY,MAINTAIN_CITY,MAINTAIN_COUNTY,CREATOR,CREAT_TIME,MODIFIER,MODIFY_TIME,STATEFLAG,BUILD_DATE,PROJECT_NAME,USER_NAME,PURPOSE,PROJECT,YJTZ,JSLY,JSYXJ,SQJSNF"/></request>''')
    HTTP_Header = {"Host": "10.209.199.74:8120","Content-Type": "application/x-www-form-urlencoded"}
    Respond_Body = requests.post(URL_Query_OpticalCable,data=XML_Info_Encoded,headers=HTTP_Header)
    Respond_Body = bytes(Respond_Body.text, encoding="utf-8")
    Respond_Body = lxml.etree.HTML(Respond_Body)
    List_Response_Key = Respond_Body.xpath("//fv/@k")
    List_Response_Value = Respond_Body.xpath("//fv/@v")
    Dic_Response = dict(zip(List_Response_Key,List_Response_Value))
    if (Dic_Response['A_EQUIP_ID'] == '') or (Dic_Response['A_EQUIP_TYPE'] == '') or (Dic_Response['A_ROOM_ID'] == ''):
        URL_Renew_OpticalCable = "http://10.209.199.74:8120/igisserver_osl/rest/ResourceController/resourcesUpdate?isUpdate=update&coreNamingRules=0"
        XML_Info_Encoded = "xml=" + urllib.parse.quote_plus('<xmldata mode="FibersegEditMode"><mc type="guanglanduan"><mo group="1"><fv k="A_OBJECT_TYPE" v="' + str(Dic_Response["A_OBJECT_TYPE"]) + '"/><fv k="AREA_LEVEL" v="' + str(Dic_Response["AREA_LEVEL"]) + '"/><fv k="STATUS" v="' + str(Dic_Response["STATUS"]) + '"/><fv k="Z_OBJECT_TYPE" v="' + str(Dic_Response["Z_OBJECT_TYPE"]) + '"/><fv k="SQJSNF" v="' + str(Dic_Response["SQJSNF"]) + '"/><fv k="Z_EQUIP_TYPE" v="' + str(Dic_Response["Z_EQUIP_TYPE"]) + '"/><fv k="Z_OBJECT_ID" v="' + str(Dic_Response["Z_OBJECT_ID"]) + '"/><fv k="A_EQUIP_ID" v="' + str(Para_List_CS_Data['A_Box_ID']) + '"/><fv k="RUWANG_DATE" v="' + str(Dic_Response["RUWANG_DATE"]) + '"/><fv k="A_EQUIP_TYPE" v="' + str(Para_List_CS_Data['A_Box_Type_ID']) + '"/><fv k="YJTZ" v="' + str(Dic_Response["YJTZ"]) + '"/><fv k="QUALITOR" v="' + str(Dic_Response["QUALITOR"]) + '"/><fv k="Z_EQUIP_ID" v="' + str(Dic_Response["Z_EQUIP_ID"]) + '"/><fv k="PROJECT_NAME" v="' + str(Dic_Response["PROJECT_NAME"]) + '"/><fv k="JSLY" v="' + str(Dic_Response["JSLY"]) + '"/><fv k="LAY_TYPE" v="' + str(Dic_Response["LAY_TYPE"]) + '"/><fv k="ASSENT_NO" v="' + str(Dic_Response["ASSENT_NO"]) + '"/><fv k="MAINTAIN_COUNTY" v="' + str(Dic_Response["MAINTAIN_COUNTY"]) + '"/><fv k="PROJECT_ID" v="' + str(Dic_Response["PROJECT_ID"]) + '"/><fv k="C_LENGTH" v="' + str(Dic_Response["C_LENGTH"]) + '"/><fv k="MAINTAIN_CITY" v="' + str(Dic_Response["MAINTAIN_CITY"]) + '"/><fv k="MAINT_DEP" v="' + str(Dic_Response["MAINT_DEP"]) + '"/><fv k="SYS_VERSION" v="' + str(Dic_Response["SYS_VERSION"]) + '"/><fv k="USER_NAME" v="' + str(Dic_Response["USER_NAME"]) + '"/><fv k="ALIAS" v="' + str(Dic_Response["ALIAS"]) + '"/><fv k="CJ_STATUS" v="' + str(Dic_Response["CJ_STATUS"]) + '"/><fv k="BUILDER" v="' + str(Dic_Response["BUILDER"]) + '"/><fv k="MAINT_STATE" v="' + str(Dic_Response["MAINT_STATE"]) + '"/><fv k="A_ROOM_ID" v="' + str(Para_List_CS_Data['A_ResPoint_ID']) + '"/><fv k="STUFF" v="' + str(Dic_Response["STUFF"]) + '"/><fv k="JSYXJ" v="' + str(Dic_Response["JSYXJ"]) + '"/><fv k="FIBER_TYPE" v="' + str(Dic_Response["FIBER_TYPE"]) + '"/><fv k="MAINT_MODE" v="' + str(Dic_Response["MAINT_MODE"]) + '"/><fv k="Z_ROOM_ID" v="' + str(Dic_Response["Z_ROOM_ID"]) + '"/><fv k="BUILD_DATE" v="' + str(Dic_Response["BUILD_DATE"]) + '"/><fv k="WIRE_SEG_TYPE" v="' + str(Dic_Response["WIRE_SEG_TYPE"]) + '"/><fv k="QUALITOR_PROJECT" v="' + str(Dic_Response["QUALITOR_PROJECT"]) + '"/><fv k="REMARK" v="' + str(Dic_Response["REMARK"]) + '"/><fv k="INT_ID" v="' + str(Dic_Response["INT_ID"]) + '"/><fv k="COUNTY_ID" v="' + str(Dic_Response["COUNTY_ID"]) + '"/><fv k="TASK_NAME" v="' + str(Dic_Response["TASK_NAME"]) + '"/><fv k="IS_ALTER" v="' + str(Dic_Response["IS_ALTER"]) + '"/><fv k="RES_OWNER" v="' + str(Dic_Response["RES_OWNER"]) + '"/><fv k="GT_VERSION" v="' + str(Dic_Response["GT_VERSION"]) + '"/><fv k="PRO_TASK_ID" v="' + str(Dic_Response["PRO_TASK_ID"]) + '"/><fv k="CJ_TASK_ID" v="' + str(Dic_Response["CJ_TASK_ID"]) + '"/><fv k="QUALITOR_COUNTY" v="' + str(Dic_Response["QUALITOR_COUNTY"]) + '"/><fv k="DIRECTION" v="' + str(Dic_Response["DIRECTION"]) + '"/><fv k="PRODUCE_DATE" v="' + str(Dic_Response["PRODUCE_DATE"]) + '"/><fv k="SERVICER" v="' + str(Dic_Response["SERVICER"]) + '"/><fv k="FIBER_NUM" v="' + str(Dic_Response["FIBER_NUM"]) + '"/><fv k="OPTI_CABLE_TYPE" v="' + str(Dic_Response["OPTI_CABLE_TYPE"]) + '"/><fv k="PROJECTCODE" v="' + str(Dic_Response["PROJECTCODE"]) + '"/><fv k="VENDOR" v="' + str(Dic_Response["VENDOR"]) + '"/><fv k="IS_WRONG" v="' + str(Dic_Response["IS_WRONG"]) + '"/><fv k="PHONE_NO" v="' + str(Dic_Response["PHONE_NO"]) + '"/><fv k="RELATED_IS_AREA" v="' + str(Dic_Response["RELATED_IS_AREA"]) + '"/><fv k="WRONG_INFO" v="' + str(Dic_Response["WRONG_INFO"]) + '"/><fv k="DIA" v="' + str(Dic_Response["DIA"]) + '"/><fv k="SERVICE_LEVEL" v="' + str(Dic_Response["SERVICE_LEVEL"]) + '"/><fv k="BUSINESS" v="' + str(Dic_Response["BUSINESS"]) + '"/><fv k="TUIWANG_DATE" v="' + str(Dic_Response["TUIWANG_DATE"]) + '"/><fv k="IS_COMPLETE_LAY" v="' + str(Dic_Response["IS_COMPLETE_LAY"]) + '"/><fv k="QRCODE" v="' + str(Dic_Response["QRCODE"]) + '"/><fv k="ZH_LABEL" v="' + str(Dic_Response["ZH_LABEL"]) + '"/><fv k="M_LENGTH" v="' + str(Dic_Response["M_LENGTH"]) + '"/><fv k="OWNERSHIP" v="' + str(Dic_Response["OWNERSHIP"]) + '"/><fv k="RELATED_SYSTEM" v="' + str(Dic_Response["RELATED_SYSTEM"]) + '"/><fv k="PURPOSE" v="' + str(Dic_Response["PURPOSE"]) + '"/><fv k="INDEX_IN_BRANCH" v="' + str(Dic_Response["INDEX_IN_BRANCH"]) + '"/><fv k="SEG_NO" v="' + str(Dic_Response["SEG_NO"]) + '"/><fv k="CITY_ID" v="' + str(Dic_Response["CITY_ID"]) + '"/><fv k="A_OBJECT_ID" v="' + str(Dic_Response["A_OBJECT_ID"]) + '"/><fv k="MAINTAINOR" v="' + str(Dic_Response["MAINTAINOR"]) + '"/><fv k="PROJECT" v="' + str(Dic_Response["PROJECT"]) + '"/></mo></mc></xmldata>')
        XML_Info_Encoded = XML_Info_Encoded.replace("/","%2f")
        Request_Lenth = chr(len(XML_Info_Encoded))
        HTTP_Header = {"Host": "10.209.199.74:8120","Content-Type": "application/x-www-form-urlencoded","Content-Length": Request_Lenth}
        Respond_Body = requests.post(URL_Renew_OpticalCable,data=XML_Info_Encoded,headers=HTTP_Header)
        Respond_Body = bytes(Respond_Body.text, encoding="utf-8")
        Respond_Body = lxml.etree.HTML(Respond_Body)
        Renew_State = Respond_Body.xpath("//@loaded")
        print('P12-' + Dic_Response["ZH_LABEL"]  +  " "  +  Renew_State[0])

def Main_Process(Para_File_Name):

    Generate_Local_Data(Para_File_Name)

    if P0_Data_Check:
        Generate_Local_Data(each_File_Name)
        Swimming_Pool(Query_Box_ID_ResPoint_ID_Alias, List_Box_Data)
        Query_Support_Sys_and_Cable_Sys()
        Query_Project_Code_ID()
        Swimming_Pool(Query_Support_Seg_ID_Cable_Seg_ID, List_CS_Data)
        Generate_Topology()
        Generate_FS_Data()
        Swimming_Pool(Query_ODM_ID_and_Terminarl_IDs, List_Box_Data)
        Swimming_Pool(Query_CS_Fiber_IDs, List_CS_Data)
        Generate_Termination_and_Direct_Melt_Data()
        Query_Run_Certification()
        Query_Create_Certification()
        Swimming_Pool(Query_POS_ID, List_Box_Data)
        Swimming_Pool(Query_POS_Port_IDs, List_Box_Data)
        Generate_OC_POS_Data_and_OC_Name()
        Query_Optical_Route_Sheet_ID()
        Query_Integrate_Sheet_ID()
        Query_OC_Int_ID()

        WB_obj = openpyxl.load_workbook(each_File_Name+'.xlsx')

        # List_Box_Data
        WS_obj = WB_obj.create_sheet('List_Box_Data')
        Dic_Column_Name = dict(sorted(List_Box_Data[0].items(), key = lambda item:item[0]))
        Column_Num = 0
        for key in Dic_Column_Name.keys():
            Column_Num += 1
            WS_obj.cell(row=1, column=Column_Num, value=key)
        Row_Num = 1
        for each_box_data in List_Box_Data:
            Row_Num += 1
            Column_Num = 0
            dic_Sorted_Box_Data = dict(sorted(each_box_data.items(), key = lambda item:item[0]))
            for value in dic_Sorted_Box_Data.values():
                Column_Num += 1
                WS_obj.cell(row=Row_Num, column=Column_Num, value=str(value))

        # List_CS_Data
        WS_obj = WB_obj.create_sheet('List_CS_Data')
        Dic_Column_Name = dict(sorted(List_CS_Data[0].items(), key = lambda item:item[0]))
        Column_Num = 0
        for key in Dic_Column_Name.keys():
            Column_Num += 1
            WS_obj.cell(row=1, column=Column_Num, value=key)
        Row_Num = 1
        for each_cs_data in List_CS_Data:
            Row_Num += 1
            Column_Num = 0
            dic_Sorted_CS_Data = dict(sorted(each_cs_data.items(), key = lambda item:item[0]))
            for value in dic_Sorted_CS_Data.values():
                Column_Num += 1
                WS_obj.cell(row=Row_Num, column=Column_Num, value=str(value))

        # List_OC_Data
        WS_obj = WB_obj.create_sheet('List_OC_Data')
        Dic_Column_Name = dict(sorted(List_OC_Data[0].items(), key = lambda item:item[0]))
        Column_Num = 0
        for key in Dic_Column_Name.keys():
            Column_Num += 1
            WS_obj.cell(row=1, column=Column_Num, value=key)
        Row_Num = 1
        for each_oc_data in List_OC_Data:
            Row_Num += 1
            Column_Num = 0
            dic_Sorted_OC_Data = dict(sorted(each_oc_data.items(), key = lambda item:item[0]))
            for value in dic_Sorted_OC_Data.values():
                Column_Num += 1
                WS_obj.cell(row=Row_Num, column=Column_Num, value=str(value))

        WB_obj.save(each_File_Name+'.xlsx')
        WB_obj.close()

    if (P1_Push_Box or 
        P2_Generate_Support_Segment or 
        P3_Generate_Cable_Segment or 
        P4_Cable_Lay or 
        P5_Generate_ODM or 
        P6_Generate_Tray or
        P7_Termination or
        P8_Direct_Melt or
        P9_Generate_Optical_Circuit or
        P10_Transmission_Design or
        P11_Termination or
        P12_Update_1_Fix_OCS):
        print('查询Box/ResPoint开始')
        Swimming_Pool(Query_Box_ID_ResPoint_ID_Alias, List_Box_Data)
        print('查询Box/ResPoint结束')

    if P1_Push_Box:
        print('P1-开始')
        Execute_Push_Box()
        print('P1-结束')

    if (P2_Generate_Support_Segment or 
        P3_Generate_Cable_Segment or 
        P4_Cable_Lay):
        print('查询Support_Sys_ID/Cable_Sys_ID')
        Query_Support_Sys_and_Cable_Sys()

    if (P2_Generate_Support_Segment or 
        P3_Generate_Cable_Segment or
        P9_Generate_Optical_Circuit or
        P10_Transmission_Design):
        print('查询Project_Code_ID')
        Query_Project_Code_ID()

    if (P2_Generate_Support_Segment or 
        P3_Generate_Cable_Segment):
        Query_Create_Certification()
        Query_Integrate_Sheet_ID()

    if P2_Generate_Support_Segment:
        print('P2-开始')
        Execute_Generate_Support_Segment()
        print('P2-结束')

    if P3_Generate_Cable_Segment:
        print('P3-开始')
        Execute_Generate_Cable_Segment()
        print('P3-结束')

    if (P4_Cable_Lay or
        P7_Termination or
        P8_Direct_Melt or
        P11_Termination):
        if (not ('Support_Seg_ID' in List_CS_Data[0])) or (not ('Cable_Seg_ID' in List_CS_Data[0])):
            print('查询Support_Seg_ID/Cable_Seg_ID开始')
            Swimming_Pool(Query_Support_Seg_ID_Cable_Seg_ID, List_CS_Data)
            print('查询Support_Seg_ID/Cable_Seg_ID结束')

    if P4_Cable_Lay:
        print('P4-开始')
        Swimming_Pool(Execute_Cable_Lay, List_CS_Data)
        print('P4-结束')

    if (P5_Generate_ODM or
        P6_Generate_Tray or
        P7_Termination or
        P8_Direct_Melt or
        P11_Termination):
        Generate_Topology()
        Generate_FS_Data()

    if P5_Generate_ODM:
        print('P5-开始')
        Swimming_Pool(Execute_Generate_ODM, List_Box_Data)
        print('P5-结束')

    if (P6_Generate_Tray or
        P7_Termination or
        P8_Direct_Melt or
        P11_Termination): # 单独添加托盘需要查询
        if not ('ODM_ID' in List_Box_Data[0]):
            print('查询ODM_ID开始')
            Swimming_Pool(Query_ODM_ID_and_Terminarl_IDs, List_Box_Data)
            print('查询ODM_ID结束')

    if P6_Generate_Tray:
        print('P6-开始')
        Swimming_Pool(Execute_Generate_Tray, List_Box_Data)
        print('P6-结束')

    if (P7_Termination or
        P8_Direct_Melt or
        P11_Termination):
        print('查询Cable_Fiber_ID开始')
        Swimming_Pool(Query_CS_Fiber_IDs, List_CS_Data)
        print('查询Cable_Fiber_ID结束')
        Generate_Termination_and_Direct_Melt_Data()

    if P7_Termination:
        print('P7-开始')
        Swimming_Pool(Execute_Termination, List_Box_Data)
        print('P7-结束')

    if P8_Direct_Melt:
        print('P8-开始')
        Swimming_Pool(Execute_Direct_Melt, List_Box_Data)
        print('P8-结束')

    if (P9_Generate_Optical_Circuit or
        P10_Transmission_Design):
        Query_Run_Certification()
        Query_Create_Certification()
        print('查询POS_ID开始')
        Swimming_Pool(Query_POS_ID, List_Box_Data)
        print('查询POS_ID结束')
        print('查询POS_Port_ID开始')
        Swimming_Pool(Query_POS_Port_IDs, List_Box_Data)
        print('查询POS_Port_ID结束')
        Generate_OC_POS_Data_and_OC_Name()
        Query_Optical_Route_Sheet_ID()

    if P9_Generate_Optical_Circuit:
        print('P9-开始')
        Swimming_Pool(Execute_Generate_Optical_Circut, List_OC_Data)
        print('P9-结束')

    if P10_Transmission_Design:
        print('查询OC_Int_ID开始')
        Query_OC_Int_ID()
        print('查询OC_Int_ID结束')

    if P10_Transmission_Design:
        # 处理重复通路的系统bug开始
        global List_OC_Path_Pointer
        List_OC_Path_Pointer = {}
        for each_oc_data in List_OC_Data:
            List_OC_Path_Pointer[each_oc_data['Z_Box_Name']] = 0
        # 处理重复通路的系统bug结束
        print('P10-开始')
        for each_oc_data in List_OC_Data:
            Execute_Transmission_Design(each_oc_data)
        # Swimming_Pool(Execute_Transmission_Design, List_OC_Data)
        print('P10-结束')

    if P11_Termination:
        print('P11-开始')
        Swimming_Pool(Execute_Termination_2nd, List_Box_Data)
        print('P11-结束')

    if P12_Update_1_Fix_OCS:
        print('P12-开始')
        Swimming_Pool(Excute_Update_1,List_CS_Data)
        print('P12-结束')

if __name__ == '__main__':
    for each_File_Name in File_Name:
        Main_Process(each_File_Name)

# print(sorted(List_CS_Data[10].items(), key = lambda item:item[0]))